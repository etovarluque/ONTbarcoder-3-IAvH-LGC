from __future__ import annotations
import os
import re
import datetime
import xlsxwriter
from collections import Counter
from typing import Optional
from PyQt5 import QtCore, QtGui, QtWidgets
from .shared import *
from .shared import _get_base_dir, _tr


class _FastaToolsWorker(QtCore.QThread):
    progress = QtCore.pyqtSignal(str)
    log_line = QtCore.pyqtSignal(str)
    finished = QtCore.pyqtSignal(list)
    error    = QtCore.pyqtSignal(str)

    def __init__(self, files: list, operation: str, params: dict, out_dir: str, parent=None):
        super().__init__(parent)
        self._files     = files
        self._operation = operation
        self._params    = params
        self._out_dir   = out_dir
        self._stop      = False

    def stop(self):
        self._stop = True

    @staticmethod
    def _read_fasta(path):
        import gzip
        opener = gzip.open if path.lower().endswith(".gz") else open
        records = []
        header = None
        seq_parts = []
        with opener(path, "rt", encoding="utf-8", errors="replace") as fh:
            for line in fh:
                line = line.rstrip("\r\n")
                if line.startswith(">"):
                    if header is not None:
                        records.append((header, "".join(seq_parts)))
                    header = line[1:]
                    seq_parts = []
                elif header is not None:
                    seq_parts.append(line)
            if header is not None:
                records.append((header, "".join(seq_parts)))
        return records

    @staticmethod
    def _write_fasta(records, path):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as fh:
            for hdr, seq in records:
                fh.write(f">{hdr}\n{seq}\n")

    @staticmethod
    def _stem(path):
        name = os.path.basename(path)
        for ext in (".fasta.gz", ".fas.gz", ".fa.gz", ".fna.gz",
                    ".fasta", ".fas", ".fa", ".fna"):
            if name.lower().endswith(ext):
                return name[: -len(ext)]
        return os.path.splitext(name)[0]

    def _iter_inputs(self):
        """Yield (display_name, stem, records) per input to process.
        In merge mode, concatenates all files and yields a single entry.
        When two files share the same stem, prefixes with the parent folder name."""
        if self._params.get("merge") and len(self._files) > 1:
            self.progress.emit("Merging files…")
            all_records = []
            for path in self._files:
                if self._stop:
                    return
                all_records.extend(self._read_fasta(path))
            yield f"Merged {len(self._files)} files", "merged", all_records
        else:
            stems = [self._stem(p) for p in self._files]
            dup_stems = {s for s in stems if stems.count(s) > 1}
            used_stems: set = set()
            for path, stem in zip(self._files, stems):
                if self._stop:
                    return
                name = os.path.basename(path)
                self.progress.emit(f"Processing {name}…")
                if stem in dup_stems:
                    parent = os.path.basename(os.path.dirname(os.path.abspath(path)))
                    stem = f"{parent}_{stem}"
                # Guarantee a unique output stem even when the parent folder also
                # matches, so a later file never silently overwrites an earlier one.
                base_stem = stem
                counter = 2
                while stem in used_stems:
                    stem = f"{base_stem}_{counter}"
                    counter += 1
                used_stems.add(stem)
                yield name, stem, self._read_fasta(path)

    @staticmethod
    def _cell_to_str(v) -> str:
        """Convert an Excel cell value to a string for ID matching.
        openpyxl returns integer-valued cells as float (e.g. 1234.0); turn those
        into "1234" so they match a "1234" FASTA id instead of becoming "1234_0"
        after the dot is normalized to an underscore."""
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)

    @staticmethod
    def _normalize_field(s: str) -> str:
        import unicodedata, re as _re
        s = str(s).strip()
        s = unicodedata.normalize("NFD", s)
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        s = _re.sub(r"[ .,]+", "_", s)
        s = _re.sub(r"_+", "_", s)
        return s.strip("_")

    def _run_unique(self):
        outputs = []
        for display_name, stem, records in self._iter_inputs():
            if self._stop:
                break
            seen = set()
            unique = []
            for hdr, seq in records:
                key = seq.upper()
                if key not in seen:
                    seen.add(key)
                    unique.append((hdr, seq))
            removed = len(records) - len(unique)
            out_path = os.path.join(self._out_dir, f"{stem}_unique.fasta")
            self._write_fasta(unique, out_path)
            self.log_line.emit(
                f"{display_name}\n"
                f"  Total sequences in :  {len(records)}\n"
                f"  Unique sequences   :  {len(unique)}\n"
                f"  Duplicates removed :  {removed}\n"
            )
            outputs.append(out_path)
        return outputs

    def _run_identical(self):
        outputs = []
        for display_name, stem, records in self._iter_inputs():
            if self._stop:
                break
            counts = Counter(seq.upper() for _, seq in records)
            identical = [(hdr, seq) for hdr, seq in records if counts[seq.upper()] > 1]
            dup_groups = sum(1 for c in counts.values() if c > 1)

            # FASTA output
            out_path = os.path.join(self._out_dir, f"{stem}_identical.fasta")
            self._write_fasta(identical, out_path)

            # Excel: group IDs by sequence
            seq_to_ids: dict = {}
            for hdr, seq in records:
                key = seq.upper()
                if counts[key] > 1:
                    seq_to_ids.setdefault(key, []).append(hdr)

            xlsx_path = os.path.join(self._out_dir, f"{stem}_duplicate_groups.xlsx")
            wb = xlsxwriter.Workbook(xlsx_path)
            hf = wb.add_format({"bold": True, "bg_color": "#4F81BD",
                                "font_color": "#FFFFFF", "border": 1})
            cf = wb.add_format({"border": 1, "text_wrap": True, "valign": "top"})
            nf = wb.add_format({"border": 1, "align": "center", "valign": "top"})

            ws_i = wb.add_worksheet("IDs")
            ws_i.activate()
            ws_g = wb.add_worksheet("Groups")

            for col, h in enumerate(["Group", "ID"]):
                ws_i.write(0, col, h, hf)
            ws_i.set_column(0, 0, 8)
            ws_i.set_column(1, 1, 30)

            for col, h in enumerate(["Group", "Count", "Sequence"]):
                ws_g.write(0, col, h, hf)
            ws_g.set_column(0, 1, 8)
            ws_g.set_column(2, 2, 50)

            id_row = 1
            for g_idx, (seq_key, ids) in enumerate(seq_to_ids.items(), start=1):
                ws_g.write(g_idx, 0, g_idx, nf)
                ws_g.write(g_idx, 1, len(ids), nf)
                # Excel caps cell text at 32767 chars; truncate longer sequences
                # so xlsxwriter does not raise/drop the cell silently.
                if len(seq_key) > 32767:
                    ws_g.write(g_idx, 2, seq_key[:32764] + "...", cf)
                else:
                    ws_g.write(g_idx, 2, seq_key, cf)
                for id_val in ids:
                    ws_i.write(id_row, 0, g_idx, nf)
                    ws_i.write(id_row, 1, id_val, cf)
                    id_row += 1
            wb.close()
            outputs.append(xlsx_path)

            self.log_line.emit(
                f"{display_name}\n"
                f"  Total sequences in       :  {len(records)}\n"
                f"  Identical sequences out  :  {len(identical)}\n"
                f"  Distinct duplicated seqs :  {dup_groups}\n"
                f"  Excel groups file        :  {os.path.basename(xlsx_path)}\n"
            )
            outputs.append(out_path)
        return outputs

    def _run_grep(self):
        import re
        pattern_str  = self._params.get("pattern", "")
        pattern_file = self._params.get("pattern_file", "")
        use_regex    = self._params.get("use_regex", False)

        patterns = []
        if pattern_file and os.path.isfile(pattern_file):
            with open(pattern_file, encoding="utf-8", errors="replace") as f:
                patterns = [ln.strip() for ln in f if ln.strip()]
        elif pattern_str:
            patterns = [pattern_str]

        if not patterns:
            raise ValueError("No pattern specified for grep operation.")

        mode_str = "regex" if use_regex else "plain text"
        pat_summary = patterns[0] if len(patterns) == 1 else f"{len(patterns)} patterns from file"

        if use_regex:
            compiled = [re.compile(p) for p in patterns]
            def matches(hdr):
                return any(rx.search(hdr) for rx in compiled)
        else:
            def matches(hdr):
                return any(p in hdr for p in patterns)

        outputs = []
        for display_name, stem, records in self._iter_inputs():
            if self._stop:
                break
            matched = [(hdr, seq) for hdr, seq in records if matches(hdr)]
            unmatched = len(records) - len(matched)
            out_path = os.path.join(self._out_dir, f"{stem}_grep.fasta")
            self._write_fasta(matched, out_path)
            self.log_line.emit(
                f"{display_name}\n"
                f"  Pattern ({mode_str})  :  {pat_summary}\n"
                f"  Total sequences in  :  {len(records)}\n"
                f"  Matched             :  {len(matched)}\n"
                f"  Not matched         :  {unmatched}\n"
            )
            outputs.append(out_path)
        return outputs

    def _run_append(self):
        import openpyxl
        excel_path = self._params.get("excel_file", "")
        separator  = self._params.get("separator", "|")

        if not excel_path or not os.path.isfile(excel_path):
            raise ValueError("No valid Excel file specified.")

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        lookup: dict = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            row_id = self._normalize_field(self._cell_to_str(row[0]))
            fields = [
                nv for v in row[1:]
                if v is not None and self._cell_to_str(v).strip()
                for nv in (self._normalize_field(self._cell_to_str(v)),)
                if any(c.isalnum() for c in nv)
            ]
            lookup[row_id] = fields

        excel_name = os.path.basename(excel_path)
        n_excel_ids = len(lookup)

        outputs = []
        for display_name, stem, records in self._iter_inputs():
            if self._stop:
                break
            updated_list = []
            n_updated = 0
            n_not_found = 0
            for hdr, seq in records:
                tokens = hdr.split()
                seq_id = self._normalize_field(tokens[0]) if tokens else ""
                fields = lookup.get(seq_id, []) if seq_id else []
                if fields:
                    new_hdr = hdr + separator + separator.join(fields)
                    n_updated += 1
                else:
                    new_hdr = hdr
                    n_not_found += 1
                updated_list.append((new_hdr, seq))
            out_path = os.path.join(self._out_dir, f"{stem}_append.fasta")
            self._write_fasta(updated_list, out_path)
            self.log_line.emit(
                f"{display_name}\n"
                f"  Excel file          :  {excel_name}  ({n_excel_ids} IDs)\n"
                f"  Total sequences in  :  {len(records)}\n"
                f"  IDs updated         :  {n_updated}\n"
                f"  IDs not in Excel    :  {n_not_found}\n"
            )
            outputs.append(out_path)
        return outputs

    def _run_reformat(self):
        mode      = self._params.get("mode", "linearize")
        wrap_cols = int(self._params.get("wrap_cols", 80))

        if mode == "linearize":
            suffix = "_linearized"
            mode_desc = "multi-line → single line"
        else:
            def _wrap(seq, w=wrap_cols):
                return "\n".join(seq[i: i + w] for i in range(0, len(seq), w)) if seq else ""
            suffix = "_wrapped"
            mode_desc = f"single line → multi-line  ({wrap_cols} nt/row)"

        outputs = []
        for display_name, stem, records in self._iter_inputs():
            if self._stop:
                break
            reformatted = records if mode == "linearize" else [(hdr, _wrap(seq)) for hdr, seq in records]
            out_path = os.path.join(self._out_dir, f"{stem}{suffix}.fasta")
            self._write_fasta(reformatted, out_path)
            self.log_line.emit(
                f"{display_name}\n"
                f"  Mode                :  {mode_desc}\n"
                f"  Total sequences     :  {len(records)}\n"
            )
            outputs.append(out_path)
        return outputs

    def _run_sort(self):
        separator = self._params.get("separator", "|")
        levels    = self._params.get("levels", [{"field": 1, "order": "asc"}])

        sep_desc = f'"{separator}"' if separator else "none (whole header)"
        levels_desc = "  |  ".join(
            f"Field {lv['field']} {'↑' if lv['order'] == 'asc' else '↓'}"
            for lv in levels
        )

        outputs = []
        for display_name, stem, records in self._iter_inputs():
            if self._stop:
                break
            working = list(records)
            for level in reversed(levels):
                idx     = level["field"] - 1
                reverse = level["order"] == "desc"

                def _key(rec, i=idx, sep=separator):
                    parts = rec[0].split(sep) if sep else [rec[0]]
                    val   = parts[i].strip() if i < len(parts) else ""
                    # Natural sort: split into (text, int) chunks so that
                    # "Sample_10" > "Sample_2" instead of "Sample_10" < "Sample_2".
                    return tuple(
                        int(c) if c.isdigit() else c.lower()
                        for c in re.split(r"(\d+)", val)
                    )

                working = sorted(working, key=_key, reverse=reverse)

            out_path = os.path.join(self._out_dir, f"{stem}_sorted.fasta")
            self._write_fasta(working, out_path)
            self.log_line.emit(
                f"{display_name}\n"
                f"  Total sequences  :  {len(records)}\n"
                f"  Separator        :  {sep_desc}\n"
                f"  Sort levels      :  {levels_desc}\n"
            )
            outputs.append(out_path)
        return outputs

    def _run_filter_fields(self):
        separator = self._params.get("separator", "|")
        criteria  = self._params.get("criteria", [])
        if not criteria:
            raise ValueError("No filter criteria specified.")

        # Parse each criterion into (field_index, op, value, numeric).
        # Operators: =, !=, >=, <=, >, <, contains, not contains
        parsed = []
        for c in criteria:
            idx = c.get("field", 1) - 1   # 0-based
            op  = c.get("op", "=")
            raw = c.get("value", "").strip()
            # Text operators must always compare as strings, even when the value
            # happens to parse as a number (e.g. "contains 5"). Otherwise the
            # numeric branch in _passes() has no case for contains/not contains
            # and the criterion is silently ignored.
            if op in ("contains", "not contains"):
                val     = raw.lower()
                numeric = False
            else:
                try:
                    val     = float(raw)
                    numeric = True
                except ValueError:
                    val     = raw.lower()
                    numeric = False
            parsed.append((idx, op, val, numeric))

        def _field_value(header, idx):
            parts = header.split(separator) if separator else [header]
            if idx >= len(parts):
                return None
            raw = parts[idx].strip()
            # For key=value fields (e.g. "coverage=24"), return the value part.
            if "=" in raw:
                raw = raw.split("=", 1)[1].strip()
            return raw

        def _passes(header):
            for idx, op, val, numeric in parsed:
                raw_v = _field_value(header, idx)
                if raw_v is None:
                    return False
                if numeric:
                    try:
                        fv = float(raw_v)
                    except ValueError:
                        return False
                    if op == ">="          and not (fv >= val): return False
                    if op == "<="          and not (fv <= val): return False
                    if op == ">"           and not (fv >  val): return False
                    if op == "<"           and not (fv <  val): return False
                    if op == "="           and not (fv == val): return False
                    if op == "!="          and not (fv != val): return False
                else:
                    sv = raw_v.lower()
                    if op == "="           and sv != val:          return False
                    if op == "!="          and sv == val:          return False
                    if op == "contains"    and val not in sv:      return False
                    if op == "not contains" and val in sv:         return False
            return True

        sep_desc  = f'"{separator}"' if separator else "none (whole header)"
        crit_desc = "  AND  ".join(
            f"field {i + 1} {o} {v}" for i, o, v, _ in parsed
        )
        outputs = []
        for display_name, stem, records in self._iter_inputs():
            if self._stop:
                break
            matched   = [(h, s) for h, s in records if _passes(h)]
            unmatched = len(records) - len(matched)
            out_path  = os.path.join(self._out_dir, f"{stem}_filtered.fasta")
            self._write_fasta(matched, out_path)
            self.log_line.emit(
                f"{display_name}\n"
                f"  Separator        :  {sep_desc}\n"
                f"  Criteria (AND)   :  {crit_desc}\n"
                f"  Total sequences  :  {len(records)}\n"
                f"  Passed           :  {len(matched)}\n"
                f"  Excluded         :  {unmatched}\n"
            )
            outputs.append(out_path)
        return outputs

    def run(self):
        try:
            os.makedirs(self._out_dir, exist_ok=True)
            if self._operation == "unique":
                outputs = self._run_unique()
            elif self._operation == "identical":
                outputs = self._run_identical()
            elif self._operation == "grep":
                outputs = self._run_grep()
            elif self._operation == "append":
                outputs = self._run_append()
            elif self._operation == "reformat":
                outputs = self._run_reformat()
            elif self._operation == "sort":
                outputs = self._run_sort()
            elif self._operation == "filter_fields":
                outputs = self._run_filter_fields()
            else:
                raise ValueError(f"Unknown operation: {self._operation}")
            if not self._stop:
                self.finished.emit(outputs)
        except Exception as exc:
            self.error.emit(str(exc))


class _DragDropLineEdit(QtWidgets.QLineEdit):
    """QLineEdit that accepts a single file dragged onto it."""

    _DRAG_STYLE = (
        f"QLineEdit {{ border: 2px solid {BLUE_MID}; background-color: {BLUE_LIGHT};"
        f" border-radius: 4px; }}"
    )

    def __init__(self, accepted_extensions=None, parent=None):
        super().__init__(parent)
        self._accepted_ext = [e.lower() for e in (accepted_extensions or [])]
        self.setAcceptDrops(True)

    def _is_valid(self, url):
        if not url.isLocalFile():
            return False
        if self._accepted_ext:
            return any(url.toLocalFile().lower().endswith(e) for e in self._accepted_ext)
        return True

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls() and any(self._is_valid(u) for u in event.mimeData().urls()):
            event.acceptProposedAction()
            self.setStyleSheet(self._DRAG_STYLE)
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls() and any(self._is_valid(u) for u in event.mimeData().urls()):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragLeaveEvent(self, event):
        self.setStyleSheet("")
        super().dragLeaveEvent(event)

    def dropEvent(self, event):
        self.setStyleSheet("")
        for url in event.mimeData().urls():
            if self._is_valid(url):
                self.setText(url.toLocalFile())
                event.acceptProposedAction()
                return
        event.ignore()


class FastaToolsPanel(QtWidgets.QWidget):

    def __init__(self, parent=None):
        super().__init__(parent)
        self._sort_field_count: int = 99

        outer = QtWidgets.QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QtWidgets.QFrame.NoFrame)
        inner = QtWidgets.QWidget()
        self._layout = QtWidgets.QVBoxLayout(inner)
        self._layout.setContentsMargins(20, 20, 20, 20)
        self._layout.setSpacing(16)
        scroll.setWidget(inner)
        outer.addWidget(scroll, 1)

        self._layout.addWidget(make_label("FASTA Tools", size=19, bold=True))
        desc = make_label(
            "Drag one or multiple FASTA files (.fa, .fas, .fasta) and choose an operation to run on them.",
            color=TEXT_SEC,
        )
        desc.setWordWrap(True)
        self._layout.addWidget(desc)

        self._drop = MultiDropZone()
        self._drop.filesDropped.connect(self._on_files)
        self._layout.addWidget(self._drop)

        self._merge_chk = QtWidgets.QCheckBox("Merge files before processing")
        self._merge_chk.setToolTip(
            "Checked: all files are concatenated into a single dataset\n"
            "and the operation runs on the combined sequences.\n"
            "Output is saved as a single merged file.\n\n"
            "Unchecked: each file is processed individually\n"
            "and a separate output file is generated for each one."
        )
        self._layout.addWidget(self._merge_chk)

        self._layout.addWidget(self._make_section_lbl("Operation"))

        ops_box = QtWidgets.QWidget()
        ops_layout = QtWidgets.QVBoxLayout(ops_box)
        ops_layout.setContentsMargins(0, 0, 0, 0)
        ops_layout.setSpacing(2)

        self._radio_unique    = QtWidgets.QRadioButton("Extract unique sequences")
        self._radio_identical = QtWidgets.QRadioButton("Extract identical sequences (duplicates)")
        self._radio_identical.setToolTip(
            "Extracts sequences that appear more than once.\n"
            "Generates two output files:\n"
            "  • <name>_identical.fasta — duplicated sequences\n"
            "  • <name>_duplicate_groups.xlsx — IDs grouped by shared sequence\n"
            "    · Sheet 'IDs': list of IDs per group\n"
            "    · Sheet 'Groups': group number, ID count and sequence"
        )
        self._radio_grep         = QtWidgets.QRadioButton("Extract sequences by pattern (grep)")
        self._radio_filter_fields = QtWidgets.QRadioButton("Filter sequences by header fields")
        self._radio_append       = QtWidgets.QRadioButton("Append info to headers from .xlsx file")
        self._radio_reformat     = QtWidgets.QRadioButton("Reformat sequence lines")
        self._radio_sort         = QtWidgets.QRadioButton("Sort sequences")

        self._op_group = QtWidgets.QButtonGroup(self)
        for rb in (self._radio_unique, self._radio_identical,
                   self._radio_grep, self._radio_filter_fields,
                   self._radio_append, self._radio_reformat, self._radio_sort):
            self._op_group.addButton(rb)

        self._radio_unique.setChecked(True)
        self._op_group.buttonClicked.connect(self._on_operation_changed)

        ops_layout.addWidget(self._radio_unique)
        ops_layout.addWidget(self._radio_identical)

        ops_layout.addSpacing(4)
        ops_layout.addWidget(self._radio_grep)
        self._grep_widget = QtWidgets.QWidget()
        gl = QtWidgets.QVBoxLayout(self._grep_widget)
        gl.setContentsMargins(20, 4, 0, 4)
        gl.setSpacing(8)

        self._grep_use_file_chk = QtWidgets.QCheckBox("Load patterns from file")
        self._grep_use_file_chk.stateChanged.connect(self._on_grep_source_changed)
        gl.addWidget(self._grep_use_file_chk)

        self._grep_pattern_row = QtWidgets.QWidget()
        gpr = QtWidgets.QHBoxLayout(self._grep_pattern_row)
        gpr.setContentsMargins(0, 0, 0, 0)
        gpr.setSpacing(8)
        gpr.addWidget(make_label("Pattern:", color=TEXT_SEC))
        self._grep_pattern_edit = QtWidgets.QLineEdit()
        self._grep_pattern_edit.setPlaceholderText("Enter search pattern…")
        gpr.addWidget(self._grep_pattern_edit, 1)
        gl.addWidget(self._grep_pattern_row)

        self._grep_file_row = QtWidgets.QWidget()
        gfr = QtWidgets.QHBoxLayout(self._grep_file_row)
        gfr.setContentsMargins(0, 0, 0, 0)
        gfr.setSpacing(8)
        lbl_pf = make_label("Patterns file:", color=TEXT_SEC)
        lbl_pf.setToolTip(
            "Plain text file — one pattern per line.\n"
            "Empty lines are ignored.\n"
            "Each pattern is matched against the full FASTA header.\n"
            "Enable 'Regex' to treat each line as a regular expression."
        )
        gfr.addWidget(lbl_pf)
        self._grep_file_edit = _DragDropLineEdit(accepted_extensions=[".txt"])
        self._grep_file_edit.setReadOnly(True)
        self._grep_file_edit.setPlaceholderText("No file selected… (or drag & drop)")
        self._grep_file_edit.setToolTip(
            "Plain text file — one pattern per line.\n"
            "Empty lines are ignored.\n"
            "Each pattern is matched against the full FASTA header.\n"
            "Enable 'Regex' to treat each line as a regular expression."
        )
        gfr.addWidget(self._grep_file_edit, 1)
        self._grep_file_clear = self._make_clear_btn()
        self._grep_file_clear.setVisible(False)
        self._grep_file_clear.clicked.connect(lambda: (
            self._grep_file_edit.clear(),
            self._grep_file_clear.setVisible(False),
        ))
        self._grep_file_edit.textChanged.connect(
            lambda t: self._grep_file_clear.setVisible(bool(t)))
        gfr.addWidget(self._grep_file_clear)
        self._grep_file_btn = QtWidgets.QPushButton("Browse…")
        self._grep_file_btn.setObjectName("secondary_btn")
        self._grep_file_btn.setFixedWidth(120)
        self._grep_file_btn.clicked.connect(self._browse_grep_file)
        gfr.addWidget(self._grep_file_btn)
        self._grep_file_row.hide()
        gl.addWidget(self._grep_file_row)

        _grep_file_note = make_label(
            "Plain text file · one pattern per line · empty lines ignored",
            size=14, color=TEXT_HINT,
        )
        gl.addWidget(_grep_file_note)
        self._grep_file_note = _grep_file_note
        _grep_file_note.hide()

        self._grep_regex_chk = QtWidgets.QCheckBox("Use regex (regular expression)")
        gl.addWidget(self._grep_regex_chk)
        self._grep_widget.hide()
        ops_layout.addWidget(self._grep_widget)

        ops_layout.addSpacing(4)
        ops_layout.addWidget(self._radio_filter_fields)
        self._filter_widget = QtWidgets.QWidget()
        ffw = QtWidgets.QVBoxLayout(self._filter_widget)
        ffw.setContentsMargins(20, 4, 0, 4)
        ffw.setSpacing(10)

        ff_note = make_label(
            "Split each header by a separator and filter by field position. "
            "All criteria must be satisfied (AND). "
            "Numeric operators: >=  <=  >  <  =  !=   ·   Text operators: =  !=  contains  not contains",
            color=TEXT_SEC, size=15,
        )
        ff_note.setWordWrap(True)
        ffw.addWidget(ff_note)

        # ── Separator row ────────────────────────────────────────────────────
        ff_sep_row = QtWidgets.QHBoxLayout()
        ff_sep_row.setSpacing(12)
        ff_sep_row.addWidget(make_label("Field separator:", color=TEXT_SEC))
        self._ff_sep_group = QtWidgets.QButtonGroup(self)
        self._ff_sep_pipe   = QtWidgets.QRadioButton('"|"  pipe')
        self._ff_sep_semi   = QtWidgets.QRadioButton('";"  semicolon')
        self._ff_sep_custom = QtWidgets.QRadioButton("Custom:")
        self._ff_sep_pipe.setChecked(True)
        for rb in (self._ff_sep_pipe, self._ff_sep_semi, self._ff_sep_custom):
            self._ff_sep_group.addButton(rb)
            ff_sep_row.addWidget(rb)
        self._ff_sep_custom_edit = QtWidgets.QLineEdit()
        self._ff_sep_custom_edit.setFixedWidth(64)
        self._ff_sep_custom_edit.setPlaceholderText("e.g. _")
        self._ff_sep_custom_edit.setMaxLength(5)
        self._ff_sep_custom_edit.hide()
        ff_sep_row.addWidget(self._ff_sep_custom_edit)
        ff_sep_row.addStretch()
        ffw.addLayout(ff_sep_row)
        self._ff_sep_group.buttonClicked.connect(self._on_ff_sep_changed)
        self._ff_sep_custom_edit.textChanged.connect(self._update_ff_preview)

        # ── Field preview ────────────────────────────────────────────────────
        self._ff_preview_frame = QtWidgets.QFrame()
        self._ff_preview_frame.setObjectName("ff_preview_card")
        self._ff_preview_frame.setStyleSheet(f"""
            QFrame#ff_preview_card {{
                background: {GRAY_BG};
                border: 1px solid {GRAY_LINE};
                border-radius: 8px;
            }}
        """)
        ff_pfl = QtWidgets.QVBoxLayout(self._ff_preview_frame)
        ff_pfl.setContentsMargins(12, 8, 12, 10)
        ff_pfl.setSpacing(6)
        ff_pfl.addWidget(make_label("Field preview  (first loaded sequence):", size=15, color=TEXT_SEC))
        self._ff_preview_lbl = QtWidgets.QLabel("—  no files loaded")
        self._ff_preview_lbl.setStyleSheet(
            f"font-family:'Courier New',Consolas,monospace; font-size:14px; color:{TEXT_PRI};"
            f" background:transparent;"
        )
        self._ff_preview_lbl.setWordWrap(True)
        ff_pfl.addWidget(self._ff_preview_lbl)
        ffw.addWidget(self._ff_preview_frame)

        # ── Criteria rows ────────────────────────────────────────────────────
        ffw.addWidget(make_label("Criteria (all must match):", color=TEXT_SEC))

        self._filter_criteria_container = QtWidgets.QWidget()
        self._filter_criteria_layout = QtWidgets.QVBoxLayout(self._filter_criteria_container)
        self._filter_criteria_layout.setContentsMargins(0, 0, 0, 0)
        self._filter_criteria_layout.setSpacing(6)
        ffw.addWidget(self._filter_criteria_container)

        self._filter_criterion_rows: list = []
        self._ff_field_count: int = 99
        self._add_filter_criterion()

        add_crit_btn = QtWidgets.QPushButton("+ Add criterion")
        add_crit_btn.setObjectName("secondary_btn")
        add_crit_btn.setFixedWidth(180)
        add_crit_btn.clicked.connect(self._add_filter_criterion)
        ffw.addWidget(add_crit_btn)

        self._filter_widget.hide()
        ops_layout.addWidget(self._filter_widget)

        ops_layout.addSpacing(4)
        ops_layout.addWidget(self._radio_append)
        self._append_widget = QtWidgets.QWidget()
        al = QtWidgets.QVBoxLayout(self._append_widget)
        al.setContentsMargins(20, 4, 0, 4)
        al.setSpacing(8)

        note = make_label(
            "Excel format: column 1 = sequence ID (same as FASTA header, without '>'), "
            "columns 2+ = fields to append. Cell values are normalized: "
            "leading/trailing spaces removed, spaces/dots/commas → '_', accents removed.",
            color=TEXT_SEC, size=15,
        )
        note.setWordWrap(True)
        al.addWidget(note)

        excel_row = QtWidgets.QHBoxLayout()
        excel_row.setSpacing(8)
        excel_row.addWidget(make_label("Excel file:", color=TEXT_SEC))
        self._excel_edit = _DragDropLineEdit(accepted_extensions=[".xlsx", ".xls"])
        self._excel_edit.setReadOnly(True)
        self._excel_edit.setPlaceholderText("No file selected… (or drag & drop)")
        excel_row.addWidget(self._excel_edit, 1)
        self._excel_clear = self._make_clear_btn()
        self._excel_clear.setVisible(False)
        self._excel_clear.clicked.connect(lambda: (
            self._excel_edit.clear(),
            self._excel_clear.setVisible(False),
        ))
        self._excel_edit.textChanged.connect(
            lambda t: self._excel_clear.setVisible(bool(t)))
        excel_row.addWidget(self._excel_clear)
        self._excel_btn = QtWidgets.QPushButton("Browse…")
        self._excel_btn.setObjectName("secondary_btn")
        self._excel_btn.setFixedWidth(120)
        self._excel_btn.clicked.connect(self._browse_excel)
        excel_row.addWidget(self._excel_btn)
        al.addLayout(excel_row)

        sep_row = QtWidgets.QHBoxLayout()
        sep_row.setSpacing(12)
        sep_row.addWidget(make_label("Field separator:", color=TEXT_SEC))
        self._sep_group = QtWidgets.QButtonGroup(self)
        self._sep_pipe = QtWidgets.QRadioButton('  "|"  pipe')
        self._sep_semi = QtWidgets.QRadioButton('  ";"  semicolon')
        self._sep_pipe.setChecked(True)
        self._sep_group.addButton(self._sep_pipe)
        self._sep_group.addButton(self._sep_semi)
        sep_row.addWidget(self._sep_pipe)
        sep_row.addWidget(self._sep_semi)
        sep_row.addStretch()
        al.addLayout(sep_row)

        self._append_widget.hide()
        ops_layout.addWidget(self._append_widget)

        ops_layout.addSpacing(4)
        ops_layout.addWidget(self._radio_reformat)
        self._reformat_widget = QtWidgets.QWidget()
        rfl = QtWidgets.QVBoxLayout(self._reformat_widget)
        rfl.setContentsMargins(20, 4, 0, 4)
        rfl.setSpacing(10)

        self._reformat_mode_group = QtWidgets.QButtonGroup(self)
        self._rf_radio_linearize  = QtWidgets.QRadioButton(
            "Multi-line → Single line  (linearize / unwrap)")
        self._rf_radio_wrap       = QtWidgets.QRadioButton(
            "Single line → Multi-line  (wrap / fold)")
        self._rf_radio_linearize.setChecked(True)
        self._reformat_mode_group.addButton(self._rf_radio_linearize)
        self._reformat_mode_group.addButton(self._rf_radio_wrap)
        self._reformat_mode_group.buttonClicked.connect(self._on_reformat_mode_changed)
        rfl.addWidget(self._rf_radio_linearize)
        rfl.addWidget(self._rf_radio_wrap)

        self._wrap_cols_row = QtWidgets.QWidget()
        wcr = QtWidgets.QHBoxLayout(self._wrap_cols_row)
        wcr.setContentsMargins(0, 0, 0, 0)
        wcr.setSpacing(10)
        wcr.addWidget(make_label("Nucleotides per row:", color=TEXT_SEC))
        self._wrap_cols_spin = QtWidgets.QSpinBox()
        self._wrap_cols_spin.setRange(10, 10000)
        self._wrap_cols_spin.setValue(80)
        self._wrap_cols_spin.setFixedWidth(90)
        wcr.addWidget(self._wrap_cols_spin)
        wcr.addStretch()
        self._wrap_cols_row.hide()
        rfl.addWidget(self._wrap_cols_row)

        self._reformat_widget.hide()
        ops_layout.addWidget(self._reformat_widget)

        ops_layout.addSpacing(4)
        ops_layout.addWidget(self._radio_sort)

        self._sort_widget = QtWidgets.QWidget()
        svl = QtWidgets.QVBoxLayout(self._sort_widget)
        svl.setContentsMargins(20, 4, 0, 4)
        svl.setSpacing(10)

        sep_note = make_label(
            "If headers contain a separator, each resulting field can be used as a sort key.",
            color=TEXT_SEC, size=15,
        )
        sep_note.setWordWrap(True)
        svl.addWidget(sep_note)

        sort_sep_row = QtWidgets.QHBoxLayout()
        sort_sep_row.setSpacing(12)
        sort_sep_row.addWidget(make_label("Field separator:", color=TEXT_SEC))
        self._sort_sep_group = QtWidgets.QButtonGroup(self)
        self._sort_sep_none = QtWidgets.QRadioButton("None (whole header)")
        self._sort_sep_pipe = QtWidgets.QRadioButton('"|"  pipe')
        self._sort_sep_semi = QtWidgets.QRadioButton('";"  semicolon')
        self._sort_sep_custom = QtWidgets.QRadioButton("Custom:")
        self._sort_sep_none.setChecked(True)
        for rb in (self._sort_sep_none, self._sort_sep_pipe, self._sort_sep_semi, self._sort_sep_custom):
            self._sort_sep_group.addButton(rb)
            sort_sep_row.addWidget(rb)
        self._sort_sep_custom_edit = QtWidgets.QLineEdit()
        self._sort_sep_custom_edit.setFixedWidth(64)
        self._sort_sep_custom_edit.setPlaceholderText("e.g. _")
        self._sort_sep_custom_edit.setMaxLength(5)
        self._sort_sep_custom_edit.hide()
        sort_sep_row.addWidget(self._sort_sep_custom_edit)
        sort_sep_row.addStretch()
        svl.addLayout(sort_sep_row)
        self._sort_sep_group.buttonClicked.connect(self._on_sort_sep_changed)
        self._sort_sep_custom_edit.textChanged.connect(self._update_sort_preview)

        self._sort_preview_frame = QtWidgets.QFrame()
        self._sort_preview_frame.setObjectName("sort_preview_card")
        self._sort_preview_frame.setStyleSheet(f"""
            QFrame#sort_preview_card {{
                background: {GRAY_BG};
                border: 1px solid {GRAY_LINE};
                border-radius: 8px;
            }}
        """)
        pfl = QtWidgets.QVBoxLayout(self._sort_preview_frame)
        pfl.setContentsMargins(12, 8, 12, 10)
        pfl.setSpacing(6)
        pfl.addWidget(make_label("Field preview  (first loaded sequence):", size=15, color=TEXT_SEC))
        self._sort_preview_lbl = QtWidgets.QLabel("—  no files loaded")
        self._sort_preview_lbl.setStyleSheet(
            f"font-family:'Courier New',Consolas,monospace; font-size:14px; color:{TEXT_PRI};"
            f" background:transparent;"
        )
        self._sort_preview_lbl.setWordWrap(True)
        pfl.addWidget(self._sort_preview_lbl)
        svl.addWidget(self._sort_preview_frame)

        svl.addWidget(make_label("Sort levels (applied top to bottom):", color=TEXT_SEC))

        self._sort_levels_container = QtWidgets.QWidget()
        self._sort_levels_layout = QtWidgets.QVBoxLayout(self._sort_levels_container)
        self._sort_levels_layout.setContentsMargins(0, 0, 0, 0)
        self._sort_levels_layout.setSpacing(6)
        svl.addWidget(self._sort_levels_container)

        self._sort_level_rows: list = []
        self._add_sort_level()

        add_level_btn = QtWidgets.QPushButton("+ Add sort level")
        add_level_btn.setObjectName("secondary_btn")
        add_level_btn.setFixedWidth(180)
        add_level_btn.clicked.connect(self._add_sort_level)
        svl.addWidget(add_level_btn)

        self._sort_widget.hide()
        ops_layout.addWidget(self._sort_widget)

        self._layout.addWidget(ops_box)

        self._status_lbl = make_label("", color=TEXT_SEC)
        self._layout.addWidget(self._status_lbl)

        self._progress_bar = QtWidgets.QProgressBar()
        self._progress_bar.setRange(0, 0)
        self._progress_bar.setFixedHeight(6)
        self._progress_bar.hide()
        self._layout.addWidget(self._progress_bar)

        self._layout.addStretch()

        self._log_edit = QtWidgets.QPlainTextEdit()
        self._log_edit.setReadOnly(True)
        self._log_edit.setFixedHeight(110)
        self._log_edit.setStyleSheet(f"""
            QPlainTextEdit {{
                background: {GRAY_BG};
                border: 1px solid {GRAY_LINE};
                border-top: 1px solid {GRAY_LINE};
                border-radius: 0px;
                font-family: 'Courier New', Consolas, monospace;
                font-size: 16px;
                color: {TEXT_PRI};
                padding: 6px;
            }}
        """)
        self._log_edit.hide()
        outer.addWidget(self._log_edit)

        footer = QtWidgets.QWidget()
        footer.setObjectName("fasta_tools_footer")
        footer.setStyleSheet(f"""
            QWidget#fasta_tools_footer {{
                background: {GRAY_CARD};
                border-top: 1px solid {GRAY_LINE};
            }}
        """)
        fl = QtWidgets.QHBoxLayout(footer)
        fl.setContentsMargins(20, 10, 20, 10)
        fl.setSpacing(8)

        self._clear_btn = QtWidgets.QPushButton("Clear")
        self._clear_btn.setObjectName("danger_btn")
        self._clear_btn.setFixedHeight(44)
        self._clear_btn.setFixedWidth(140)
        self._clear_btn.clicked.connect(self._clear)
        fl.addWidget(self._clear_btn)

        self._open_folder_btn = QtWidgets.QPushButton("Open folder  📂")
        self._open_folder_btn.setObjectName("secondary_btn")
        self._open_folder_btn.setFixedHeight(44)
        self._open_folder_btn.hide()
        self._open_folder_btn.clicked.connect(self._open_output_folder)
        fl.addWidget(self._open_folder_btn)
        fl.addStretch()

        self._run_btn = QtWidgets.QPushButton("Run  →")
        self._run_btn.setObjectName("primary_btn")
        self._run_btn.setFixedHeight(44)
        self._run_btn.setFixedWidth(300)
        self._run_btn.clicked.connect(self._on_run_clicked)
        self._set_run_enabled(False)
        fl.addWidget(self._run_btn)
        outer.addWidget(footer)

        self._worker: Optional[_FastaToolsWorker] = None
        self._files: list = []
        self._last_outputs: list = []

    @staticmethod
    def _make_section_lbl(text):
        lbl = make_section_label(text)
        lbl.setStyleSheet(
            f"font-size:17px; font-weight:600; color:{BLUE}; letter-spacing:0.4px;"
        )
        return lbl

    @staticmethod
    def _make_clear_btn():
        btn = QtWidgets.QPushButton("✕")
        btn.setFixedSize(26, 26)
        btn.setToolTip("Clear")
        btn.setStyleSheet(
            f"QPushButton {{ background-color: transparent; color: {TEXT_HINT};"
            f" border: none; border-radius: 5px; font-size: 13px; font-weight: bold; }}"
            f"QPushButton:hover {{ background-color: {RED_LT}; color: {RED}; }}"
            f"QPushButton:pressed {{ background-color: {RED}; color: white; }}"
        )
        return btn

    def _set_run_enabled(self, enabled: bool):
        self._run_btn.setEnabled(enabled)
        if enabled:
            self._run_btn.setStyleSheet(
                f"QPushButton {{ background-color:{BLUE}; color:white; border:none; "
                f"border-radius:8px; padding:9px 20px; font-size:18px; font-weight:500; }}"
                f"QPushButton:hover {{ background-color:#0C4A82; }}"
            )
        else:
            self._run_btn.setStyleSheet(
                f"QPushButton {{ background-color:{GRAY_LINE}; color:{TEXT_HINT}; border:none; "
                f"border-radius:8px; padding:9px 20px; font-size:18px; font-weight:500; }}"
            )

    def _on_files(self, files: list):
        self._files = files
        self._set_run_enabled(bool(files))
        self._status_lbl.setText("")
        self._status_lbl.setStyleSheet("")
        self._update_sort_preview()
        self._update_ff_preview()

    def _on_operation_changed(self, _btn):
        self._grep_widget.setVisible(self._radio_grep.isChecked())
        self._filter_widget.setVisible(self._radio_filter_fields.isChecked())
        self._append_widget.setVisible(self._radio_append.isChecked())
        self._reformat_widget.setVisible(self._radio_reformat.isChecked())
        self._sort_widget.setVisible(self._radio_sort.isChecked())
        if self._radio_sort.isChecked():
            self._update_sort_preview()
        if self._radio_filter_fields.isChecked():
            self._update_ff_preview()
        self._log_edit.clear()
        self._log_edit.hide()
        self._status_lbl.setText("")
        self._status_lbl.setStyleSheet("")
        self._open_folder_btn.hide()

    def _on_reformat_mode_changed(self, _btn):
        self._wrap_cols_row.setVisible(self._rf_radio_wrap.isChecked())

    def _on_sort_sep_changed(self, *_):
        self._sort_sep_custom_edit.setVisible(self._sort_sep_custom.isChecked())
        self._update_sort_preview()

    def _update_sort_preview(self, *_):
        if not self._files:
            self._sort_preview_lbl.setText("—  no files loaded")
            return
        path = self._files[0]
        try:
            import gzip as _gz
            opener = _gz.open if path.lower().endswith(".gz") else open
            header = ""
            with opener(path, "rt", encoding="utf-8", errors="replace") as fh:
                for line in fh:
                    if line.startswith(">"):
                        header = line[1:].rstrip()
                        break
        except Exception:
            self._sort_preview_lbl.setText("—  could not read file")
            return
        if not header:
            self._sort_preview_lbl.setText("—  no sequences found")
            return
        sep = ("|" if self._sort_sep_pipe.isChecked()
               else ";" if self._sort_sep_semi.isChecked()
               else self._sort_sep_custom_edit.text() if self._sort_sep_custom.isChecked()
               else "")
        # Match the worker's split (unconditional when a separator is set), so the
        # field numbering in the preview always lines up with the actual sort.
        parts = header.split(sep) if sep else [header]
        self._sort_preview_lbl.setText(
            "\n".join(f"Field {i:>2}:  {p.strip()}" for i, p in enumerate(parts, 1))
        )
        self._sort_field_count = len(parts)
        self._update_field_spin_limits()

    def _update_field_spin_limits(self):
        for row in self._sort_level_rows:
            field_spin = row[1]
            field_spin.setMaximum(self._sort_field_count)
            if field_spin.value() > self._sort_field_count:
                field_spin.setValue(self._sort_field_count)

    def _add_sort_level(self):
        n = len(self._sort_level_rows) + 1

        row_w = QtWidgets.QWidget()
        rl = QtWidgets.QHBoxLayout(row_w)
        rl.setContentsMargins(0, 0, 0, 0)
        rl.setSpacing(8)

        lbl = make_label(f"Level {n}:", color=TEXT_SEC)
        lbl.setFixedWidth(60)
        rl.addWidget(lbl)

        rl.addWidget(make_label("Field:", color=TEXT_SEC))
        field_spin = QtWidgets.QSpinBox()
        field_spin.setRange(1, self._sort_field_count)
        field_spin.setValue(min(n, self._sort_field_count))
        field_spin.setFixedWidth(70)
        rl.addWidget(field_spin)

        rl.addSpacing(8)
        rl.addWidget(make_label("Order:", color=TEXT_SEC))
        order_combo = QtWidgets.QComboBox()
        order_combo.addItem("Ascending  ↑", "asc")
        order_combo.addItem("Descending ↓", "desc")
        order_combo.setFixedWidth(160)
        rl.addWidget(order_combo)

        if n > 1:
            remove_btn = QtWidgets.QPushButton("✕")
            remove_btn.setFixedSize(28, 28)
            remove_btn.setStyleSheet(
                "QPushButton { background:transparent; border:1px solid #CCC;"
                " border-radius:4px; color:#888; font-size:11px; }"
                "QPushButton:hover { background:#FEE2E2; border-color:#EF4444; color:#EF4444; }"
            )
            remove_btn.clicked.connect(lambda _, w=row_w: self._remove_sort_level(w))
            rl.addWidget(remove_btn)

        rl.addStretch()
        self._sort_levels_layout.addWidget(row_w)
        self._sort_level_rows.append((row_w, field_spin, order_combo))

    def _remove_sort_level(self, row_widget):
        for i, (w, _fs, _oc) in enumerate(self._sort_level_rows):
            if w is row_widget:
                self._sort_level_rows.pop(i)
                self._sort_levels_layout.removeWidget(w)
                w.deleteLater()
                self._renumber_sort_levels()
                break

    def _renumber_sort_levels(self):
        for i, (w, _fs, _oc) in enumerate(self._sort_level_rows):
            lbl = w.layout().itemAt(0).widget()
            if isinstance(lbl, QtWidgets.QLabel):
                lbl.setText(f"Level {i + 1}:")

    def _on_ff_sep_changed(self, *_):
        self._ff_sep_custom_edit.setVisible(self._ff_sep_custom.isChecked())
        self._update_ff_preview()

    def _update_ff_preview(self, *_):
        if not self._files:
            self._ff_preview_lbl.setText("—  no files loaded")
            return
        path = self._files[0]
        try:
            import gzip as _gz
            opener = _gz.open if path.lower().endswith(".gz") else open
            header = ""
            with opener(path, "rt", encoding="utf-8", errors="replace") as fh:
                for line in fh:
                    if line.startswith(">"):
                        header = line[1:].rstrip()
                        break
        except Exception:
            self._ff_preview_lbl.setText("—  could not read file")
            return
        if not header:
            self._ff_preview_lbl.setText("—  no sequences found")
            return
        sep = ("|" if self._ff_sep_pipe.isChecked()
               else ";" if self._ff_sep_semi.isChecked()
               else self._ff_sep_custom_edit.text() if self._ff_sep_custom.isChecked()
               else "|")
        # Match the worker's split (unconditional when a separator is set), so the
        # field numbering in the preview always lines up with the actual filter.
        parts = header.split(sep) if sep else [header]
        self._ff_preview_lbl.setText(
            "\n".join(f"Field {i:>2}:  {p.strip()}" for i, p in enumerate(parts, 1))
        )
        self._ff_field_count = len(parts)
        self._update_ff_spin_limits()

    def _update_ff_spin_limits(self):
        for _w, fs, _oc, _ve in self._filter_criterion_rows:
            fs.setMaximum(self._ff_field_count)
            if fs.value() > self._ff_field_count:
                fs.setValue(self._ff_field_count)

    def _add_filter_criterion(self):
        _REMOVE_STYLE = (
            "QPushButton { background:transparent; border:1px solid #CCC;"
            " border-radius:4px; color:#888; font-size:11px; }"
            "QPushButton:hover { background:#FEE2E2; border-color:#EF4444; color:#EF4444; }"
        )
        n = len(self._filter_criterion_rows) + 1
        row_w = QtWidgets.QWidget()
        rl = QtWidgets.QHBoxLayout(row_w)
        rl.setContentsMargins(0, 0, 0, 0)
        rl.setSpacing(6)

        rl.addWidget(make_label("Field:", color=TEXT_SEC))
        field_spin = QtWidgets.QSpinBox()
        field_spin.setRange(1, self._ff_field_count)
        field_spin.setValue(min(n, self._ff_field_count))
        field_spin.setFixedWidth(70)
        rl.addWidget(field_spin)

        op_combo = QtWidgets.QComboBox()
        for op in ("=", "!=", ">=", "<=", ">", "<", "contains", "not contains"):
            op_combo.addItem(op)
        op_combo.setFixedWidth(130)
        rl.addWidget(op_combo)

        value_edit = QtWidgets.QLineEdit()
        value_edit.setPlaceholderText("value")
        value_edit.setFixedWidth(140)
        rl.addWidget(value_edit)

        if self._filter_criterion_rows:
            remove_btn = QtWidgets.QPushButton("✕")
            remove_btn.setFixedSize(28, 28)
            remove_btn.setStyleSheet(_REMOVE_STYLE)
            remove_btn.clicked.connect(lambda _, w=row_w: self._remove_filter_criterion(w))
            rl.addWidget(remove_btn)

        rl.addStretch()
        self._filter_criteria_layout.addWidget(row_w)
        self._filter_criterion_rows.append((row_w, field_spin, op_combo, value_edit))

    def _remove_filter_criterion(self, row_widget):
        for i, (w, _fs, _oc, _ve) in enumerate(self._filter_criterion_rows):
            if w is row_widget:
                self._filter_criterion_rows.pop(i)
                self._filter_criteria_layout.removeWidget(w)
                w.deleteLater()
                break

    def _on_grep_source_changed(self, state):
        use_file = (state == QtCore.Qt.Checked)
        self._grep_pattern_row.setVisible(not use_file)
        self._grep_file_row.setVisible(use_file)
        self._grep_file_note.setVisible(use_file)

    def _browse_grep_file(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Select patterns file", "", "Text files (*.txt);;All files (*)"
        )
        if path:
            self._grep_file_edit.setText(path)

    def _browse_excel(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Select Excel file", "", "Excel files (*.xlsx *.xls)"
        )
        if path:
            self._excel_edit.setText(path)

    def _clear(self):
        if self._worker and self._worker.isRunning():
            for sig, slot in (
                (self._worker.progress, self._on_progress),
                (self._worker.log_line, self._on_log_line),
                (self._worker.finished, self._on_finished),
                (self._worker.error,    self._on_error),
            ):
                try:
                    sig.disconnect(slot)
                except RuntimeError:
                    pass
            self._worker.stop()
            self._worker.wait(500)

        self._drop.clear()
        self._files = []
        self._last_outputs = []
        self._status_lbl.setText("")
        self._status_lbl.setStyleSheet("")
        self._progress_bar.hide()
        self._log_edit.clear()
        self._log_edit.hide()
        self._open_folder_btn.hide()
        self._set_run_enabled(False)
        self._merge_chk.setChecked(False)

        self._radio_unique.setChecked(True)
        self._grep_widget.hide()
        self._filter_widget.hide()
        self._append_widget.hide()
        self._reformat_widget.hide()
        self._sort_widget.hide()

        self._grep_use_file_chk.setChecked(False)
        self._grep_pattern_edit.clear()
        self._grep_file_edit.clear()
        self._grep_regex_chk.setChecked(False)
        self._grep_pattern_row.show()
        self._grep_file_row.hide()
        self._grep_file_note.hide()

        self._excel_edit.clear()
        self._sep_pipe.setChecked(True)

        self._rf_radio_linearize.setChecked(True)
        self._wrap_cols_spin.setValue(80)
        self._wrap_cols_row.hide()

        self._sort_sep_none.setChecked(True)
        self._sort_sep_custom_edit.clear()
        self._sort_sep_custom_edit.hide()
        while len(self._sort_level_rows) > 1:
            w, _fs, _oc = self._sort_level_rows[-1]
            self._sort_level_rows.pop()
            self._sort_levels_layout.removeWidget(w)
            w.deleteLater()
        if self._sort_level_rows:
            _w, fs, oc = self._sort_level_rows[0]
            fs.setValue(1)
            oc.setCurrentIndex(0)
        self._update_sort_preview()

        self._ff_sep_pipe.setChecked(True)
        self._ff_sep_custom_edit.clear()
        self._ff_sep_custom_edit.hide()
        while self._filter_criterion_rows:
            w, _fs, _oc, _ve = self._filter_criterion_rows.pop()
            self._filter_criteria_layout.removeWidget(w)
            w.deleteLater()
        self._add_filter_criterion()

    def _on_run_clicked(self):
        if not self._files:
            return

        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        folder_name = f"ont-barcoder_{ts}_fastools"
        auto_dir = os.path.join(_get_base_dir(), "output", folder_name)

        out_dir = self._ask_output_dir(auto_dir, folder_name)
        if out_dir is None:
            return

        operation = (
            "unique"         if self._radio_unique.isChecked()         else
            "identical"      if self._radio_identical.isChecked()      else
            "grep"           if self._radio_grep.isChecked()           else
            "filter_fields"  if self._radio_filter_fields.isChecked()  else
            "append"         if self._radio_append.isChecked()         else
            "reformat"       if self._radio_reformat.isChecked()       else
            "sort"
        )

        params: dict = {"merge": self._merge_chk.isChecked()}
        if operation == "grep":
            params["use_regex"] = self._grep_regex_chk.isChecked()
            if self._grep_use_file_chk.isChecked():
                params["pattern_file"] = self._grep_file_edit.text().strip()
            else:
                params["pattern"] = self._grep_pattern_edit.text().strip()
        elif operation == "filter_fields":
            if self._ff_sep_semi.isChecked():
                params["separator"] = ";"
            elif self._ff_sep_custom.isChecked():
                custom_sep = self._ff_sep_custom_edit.text()
                if not custom_sep:
                    QtWidgets.QMessageBox.warning(
                        self, "Filter configuration",
                        "Custom separator is empty.\n"
                        "Please enter a separator character or choose a different option."
                    )
                    return
                params["separator"] = custom_sep
            else:
                params["separator"] = "|"
            criteria = [
                {"field": fs.value(), "op": oc.currentText(), "value": ve.text().strip()}
                for _w, fs, oc, ve in self._filter_criterion_rows
                if ve.text().strip()
            ]
            if not criteria:
                QtWidgets.QMessageBox.warning(
                    self, "Filter configuration",
                    "No valid criteria.\n"
                    "Please fill in at least one value."
                )
                return
            # Numeric operators need a numeric value; otherwise the criterion
            # would fall into the text branch (which has no case for them) and
            # silently pass every sequence.
            for c in criteria:
                if c["op"] in (">=", "<=", ">", "<"):
                    try:
                        float(c["value"])
                    except ValueError:
                        QtWidgets.QMessageBox.warning(
                            self, "Filter configuration",
                            f"Operator '{c['op']}' (field {c['field']}) requires a "
                            f"numeric value, but got “{c['value']}”.\n\n"
                            "Enter a number or use a text operator "
                            "(=, !=, contains, not contains)."
                        )
                        return
            params["criteria"] = criteria
        elif operation == "append":
            params["excel_file"] = self._excel_edit.text().strip()
            params["separator"]  = "|" if self._sep_pipe.isChecked() else ";"
        elif operation == "reformat":
            params["mode"]      = "wrap" if self._rf_radio_wrap.isChecked() else "linearize"
            params["wrap_cols"] = self._wrap_cols_spin.value()
        elif operation == "sort":
            if self._sort_sep_pipe.isChecked():
                params["separator"] = "|"
            elif self._sort_sep_semi.isChecked():
                params["separator"] = ";"
            elif self._sort_sep_custom.isChecked():
                custom_sep = self._sort_sep_custom_edit.text()
                if not custom_sep:
                    QtWidgets.QMessageBox.warning(
                        self, "Sort configuration",
                        "Custom separator is empty.\n"
                        "Please enter a separator character or choose a different option."
                    )
                    return
                params["separator"] = custom_sep
            else:
                params["separator"] = ""
            params["levels"] = [
                {"field": fs.value(), "order": oc.currentData()}
                for _, fs, oc in self._sort_level_rows
            ]
            if not params["separator"]:
                bad = [lv for lv in params["levels"] if lv["field"] > 1]
                if bad:
                    QtWidgets.QMessageBox.warning(
                        self, "Sort configuration",
                        "Separator is set to None (whole header = 1 field).\n"
                        "All sort levels must use Field 1.\n\n"
                        "Please select a separator or set all levels to Field 1."
                    )
                    return

        self._progress_bar.show()
        self._status_lbl.setStyleSheet("")
        self._status_lbl.setText("Running…")
        self._set_run_enabled(False)
        self._log_edit.clear()
        self._log_edit.show()

        self._worker = _FastaToolsWorker(self._files, operation, params, out_dir)
        self._worker.progress.connect(self._on_progress)
        self._worker.log_line.connect(self._on_log_line)
        self._worker.finished.connect(self._on_finished)
        self._worker.error.connect(self._on_error)
        self._worker.start()

    def _ask_output_dir(self, auto_dir: str, folder_name: str) -> Optional[str]:
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Output folder")
        dlg.setMinimumWidth(480)
        dlg.setStyleSheet(f"""
            QDialog {{ background-color: {GRAY_CARD}; }}
            QLabel {{ color: {TEXT_PRI}; background-color: transparent; }}
            QRadioButton {{
                color: {TEXT_PRI}; background-color: transparent;
                font-size: 15px; padding: 6px 0;
            }}
            QRadioButton::indicator {{ width: 16px; height: 16px; }}
            QPushButton {{
                border-radius: 8px; padding: 8px 20px;
                font-size: 15px; font-weight: 500;
            }}
            #dlg_ok_btn {{ background-color: {BLUE}; color: white; border: none; }}
            #dlg_ok_btn:hover {{ background-color: #0C4A82; }}
            #dlg_cancel_btn {{
                background-color: transparent; color: {BLUE};
                border: 1px solid {BLUE};
            }}
            #dlg_cancel_btn:hover {{ background-color: {BLUE_LIGHT}; }}
        """)

        vlay = QtWidgets.QVBoxLayout(dlg)
        vlay.setSpacing(16)
        vlay.setContentsMargins(24, 24, 24, 20)

        title_lbl = QtWidgets.QLabel("Where to save the results?")
        title_lbl.setStyleSheet(f"font-size:17px; font-weight:700; color:{TEXT_PRI};")
        vlay.addWidget(title_lbl)

        radio_auto = QtWidgets.QRadioButton(
            f"Automatic folder (recommended)\n  …/output/{folder_name}/"
        )
        radio_auto.setChecked(True)
        radio_custom = QtWidgets.QRadioButton("Select folder manually")
        vlay.addWidget(radio_auto)
        vlay.addWidget(radio_custom)
        vlay.addSpacing(8)

        btn_row = QtWidgets.QHBoxLayout()
        btn_row.addStretch()
        btn_cancel = QtWidgets.QPushButton("Cancel")
        btn_cancel.setObjectName("dlg_cancel_btn")
        btn_cancel.setFixedHeight(38)
        btn_ok = QtWidgets.QPushButton("Run")
        btn_ok.setObjectName("dlg_ok_btn")
        btn_ok.setFixedHeight(38)
        btn_ok.setDefault(True)
        btn_cancel.clicked.connect(dlg.reject)
        btn_ok.clicked.connect(dlg.accept)
        btn_row.addWidget(btn_cancel)
        btn_row.addSpacing(8)
        btn_row.addWidget(btn_ok)
        vlay.addLayout(btn_row)

        if dlg.exec_() != QtWidgets.QDialog.Accepted:
            return None

        if radio_auto.isChecked():
            return auto_dir
        parent_dir = QtWidgets.QFileDialog.getExistingDirectory(
            self, "Select output folder"
        )
        if not parent_dir:
            return None
        return os.path.join(parent_dir, folder_name)

    def _on_progress(self, msg: str):
        self._status_lbl.setText(msg)

    def _on_log_line(self, line: str):
        self._log_edit.appendPlainText(line)

    def _on_finished(self, outputs: list):
        self._progress_bar.hide()
        self._set_run_enabled(True)
        self._last_outputs = outputs
        self._status_lbl.setStyleSheet(f"color:{GREEN};")
        self._status_lbl.setText(f"Done — {len(outputs)} file(s) saved.")
        self._log_edit.appendPlainText(
            f"─── {len(outputs)} file(s) saved to: "
            f"{os.path.dirname(outputs[0]) if outputs else '—'}"
        )
        if outputs:
            self._open_folder_btn.show()

    def _on_error(self, msg: str):
        self._progress_bar.hide()
        self._set_run_enabled(True)
        self._status_lbl.setStyleSheet(f"color:{RED};")
        self._status_lbl.setText(f"Error: {msg}")
        self._log_edit.appendPlainText(f"ERROR: {msg}")

    def _open_output_folder(self):
        if not self._last_outputs:
            return
        folder = os.path.dirname(self._last_outputs[0])
        if os.path.isdir(folder):
            QtGui.QDesktopServices.openUrl(
                QtCore.QUrl.fromLocalFile(folder)
            )
