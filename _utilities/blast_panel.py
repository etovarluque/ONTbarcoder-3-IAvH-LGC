from __future__ import annotations
import os
import time
import datetime
import threading
from typing import Dict, List, Optional, Tuple
from PyQt5 import QtCore, QtGui, QtWidgets
from .shared import *
from .shared import _get_base_dir, _profiles_dir, _tr, _json_mod


# ═══════════════════════════════════════════════════════════════════════════
# BLAST PANEL
# ═══════════════════════════════════════════════════════════════════════════

class BlastPanel(QtWidgets.QWidget):
    blastRequested = QtCore.pyqtSignal(list, dict)   # files, config dict
    stopRequested  = QtCore.pyqtSignal()             # user clicked Stop

    _DATABASES        = ["core_nt", "nt", "refseq_rna", "16S_ribosomal_RNA"]
    _PROGRAMS         = ["blastn&MEGABLAST=on", "blastn", "megablast"]
    _PROGRAM_LABELS   = ["blastn + MEGABLAST (recommended)", "blastn", "megablast"]

    # Live-log slot keys (fixed lines, updated in place)
    _SLOT_KEYS = ("info", "blast", "organism", "taxonomy", "progress", "result")

    def __init__(self, parent=None):
        super().__init__(parent)

        outer_layout = QtWidgets.QVBoxLayout(self)
        outer_layout.setContentsMargins(0, 0, 0, 0)
        outer_layout.setSpacing(0)

        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QtWidgets.QFrame.NoFrame)

        self._inner = QtWidgets.QWidget()
        self._layout = QtWidgets.QVBoxLayout(self._inner)
        self._layout.setContentsMargins(20, 20, 20, 8)
        self._layout.setSpacing(14)
        scroll.setWidget(self._inner)
        outer_layout.addWidget(scroll, 0)   # scroll takes only what content needs

        # ── Title + description ──
        self._lbl_title = make_label("BLAST NCBI Search", size=19, bold=True)
        self._lbl_desc = make_label(
            "BLAST sequences in NCBI. "
            "Drag-and-drop one or more FASTA files (.fa, .fas, .fasta).\n"
            "Optional: results might include organism and taxonomic classification from NCBI Taxonomy.",
            color=TEXT_SEC
        )
        self._lbl_desc.setWordWrap(True)
        self._layout.addWidget(self._lbl_title)
        self._layout.addWidget(self._lbl_desc)

        # ── Settings group ──
        self._settings_box = QtWidgets.QGroupBox("BLAST Settings")
        self._settings_box.setStyleSheet("QGroupBox { font-weight:600; color:#1A1A2E; }")
        sg = QtWidgets.QFormLayout(self._settings_box)
        sg.setLabelAlignment(QtCore.Qt.AlignRight)
        sg.setSpacing(10)
        sg.setContentsMargins(16, 16, 16, 16)

        self._lbl_api = QtWidgets.QLabel("NCBI API Key:")
        api_row = QtWidgets.QWidget()
        al = QtWidgets.QHBoxLayout(api_row)
        al.setContentsMargins(0, 0, 0, 0)
        al.setSpacing(4)
        self._api_key_edit = QtWidgets.QLineEdit()
        self._api_key_edit.setPlaceholderText("Paste your NCBI API key here…")
        self._api_key_edit.setMinimumWidth(500)
        self._api_key_edit.setText(self._load_api_key())
        self._api_key_edit.editingFinished.connect(self._save_api_key)
        self._api_key_edit.textChanged.connect(self._on_api_key_changed)
        al.addWidget(self._api_key_edit)
        self._api_key_clear_btn = QtWidgets.QPushButton("✕")
        self._api_key_clear_btn.setFixedSize(28, 28)
        self._api_key_clear_btn.setToolTip("Clear API key")
        self._api_key_clear_btn.setStyleSheet(
            "QPushButton { background: transparent; border: 1px solid #CCC;"
            " border-radius: 4px; color: #888; font-size:11px; }"
            "QPushButton:hover { background: #FEE2E2; border-color: #EF4444; color: #EF4444; }"
        )
        self._api_key_clear_btn.clicked.connect(self._clear_api_key)
        al.addWidget(self._api_key_clear_btn)
        al.addStretch()
        sg.addRow(self._lbl_api, api_row)

        self._lbl_api_warn = QtWidgets.QLabel(
            "<table cellspacing='0' cellpadding='0'><tr>"
            "<td valign='top'>⚠&nbsp;&nbsp;</td>"
            "<td>A personal NCBI API key is required — key is saved automatically.<br>"
            "Register free at: "
            "<a href='https://www.ncbi.nlm.nih.gov/account' "
            "style='color:#185FA5;'>ncbi.nlm.nih.gov/account</a></td>"
            "</tr></table>"
        )
        self._lbl_api_warn.setWordWrap(True)
        self._lbl_api_warn.setOpenExternalLinks(True)
        self._lbl_api_warn.setStyleSheet("color:#B45309; font-size:16px;")
        sg.addRow("", self._lbl_api_warn)
        self._lbl_api_warn.setVisible(not bool(self._api_key_edit.text().strip()))

        self._db_combo = QtWidgets.QComboBox()
        for db in self._DATABASES:
            self._db_combo.addItem(db)
        self._db_combo.setFixedWidth(500)
        self._lbl_db = QtWidgets.QLabel("Database:")
        sg.addRow(self._lbl_db, self._db_combo)

        self._prog_combo = QtWidgets.QComboBox()
        for lbl in self._PROGRAM_LABELS:
            self._prog_combo.addItem(lbl)
        self._prog_combo.setFixedWidth(500)
        self._lbl_prog = QtWidgets.QLabel("Program:")
        sg.addRow(self._lbl_prog, self._prog_combo)

        hits_row = QtWidgets.QWidget()
        hl = QtWidgets.QHBoxLayout(hits_row)
        hl.setContentsMargins(0, 0, 0, 0)
        self._hits_spin = QtWidgets.QSpinBox()
        self._hits_spin.setRange(1, 100)
        self._hits_spin.setValue(5)
        self._hits_spin.setFixedWidth(80)
        hl.addWidget(self._hits_spin)
        hl.addStretch()
        self._lbl_hits = QtWidgets.QLabel("Hits per sequence (1–100):")
        sg.addRow(self._lbl_hits, hits_row)

        batch_row = QtWidgets.QWidget()
        bl2 = QtWidgets.QHBoxLayout(batch_row)
        bl2.setContentsMargins(0, 0, 0, 0)
        bl2.setSpacing(6)
        self._batch_spin = QtWidgets.QSpinBox()
        self._batch_spin.setRange(1, 100)
        self._batch_spin.setValue(50)
        self._batch_spin.setFixedWidth(80)
        bl2.addWidget(self._batch_spin)
        self._ncbi_warn_icon = QtWidgets.QLabel("⚠")
        self._ncbi_warn_icon.setStyleSheet(
            "color: #B45309; font-size: 17px; padding: 0 2px;"
        )
        self._ncbi_warn_icon.setToolTip(
            "<b>NCBI usage policy warning</b><br><br>"
            "NCBI monitors and penalizes excessive use of its servers.<br>"
            "Submitting too many requests in a short period may result in:<br>"
            "• Temporary or permanent IP blocking<br>"
            "• Suspension of your API key<br><br>"
            "Keep batches ≤ 50 sequences and avoid running multiple<br>"
            "simultaneous BLAST sessions or long sessions in NCBI."
        )
        self._ncbi_warn_icon.setCursor(QtCore.Qt.WhatsThisCursor)
        bl2.addWidget(self._ncbi_warn_icon)
        bl2.addStretch()
        self._lbl_batch = QtWidgets.QLabel("Sequences per batch (max 50 recommended):")
        sg.addRow(self._lbl_batch, batch_row)

        self._tax_check = QtWidgets.QCheckBox("Fetch organism + taxonomic classification")
        self._tax_check.setChecked(True)
        self._lbl_tax = QtWidgets.QLabel("Taxonomy lookup:")
        sg.addRow(self._lbl_tax, self._tax_check)

        self._layout.addWidget(self._settings_box)

        # ── Drop zone ──
        self._drop = MultiDropZone()
        self._drop.filesDropped.connect(self._on_files)
        self._layout.addWidget(self._drop)
        self._layout.addStretch()   # packs content at top; log lives outside the scroll

        # ── Live progress display (outside scroll so it expands to fill space) ──
        self._log_slots = {k: "" for k in self._SLOT_KEYS}
        self._log = QtWidgets.QPlainTextEdit()
        self._log.setReadOnly(True)
        self._log.setFont(QtGui.QFont("Consolas", 9))
        self._log.setMinimumHeight(200)
        self._log.setSizePolicy(
            QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed
        )
        self._log.setStyleSheet(
            f"QPlainTextEdit {{ background:{GRAY_BG}; border:1px solid {GRAY_LINE}; "
            f"border-radius:6px; padding:6px; color:{TEXT_PRI}; margin:0 20px 8px 20px; "
            f"font-family:'Consolas','Courier New',monospace; }}"
        )
        self._log.hide()
        outer_layout.addWidget(self._log, 0)   # stretch=0: fixed height (~30% of panel, min 200px)

        # ── Elapsed-time timer ──
        self._info_base   = ""    # static part of the info line (without time)
        self._start_time  = 0.0
        self._elapsed_timer = QtCore.QTimer(self)
        self._elapsed_timer.setInterval(1000)
        self._elapsed_timer.timeout.connect(self._tick_elapsed)

        # ── Footer ──
        footer = QtWidgets.QWidget()
        footer.setObjectName("blast_footer")
        footer.setStyleSheet(f"""
            QWidget#blast_footer {{
                background: {GRAY_CARD};
                border-top: 1px solid {GRAY_LINE};
            }}
        """)
        fl = QtWidgets.QHBoxLayout(footer)
        fl.setContentsMargins(20, 10, 20, 10)

        self._clear_btn = QtWidgets.QPushButton("Clear")
        self._clear_btn.setObjectName("danger_btn")
        self._clear_btn.setFixedHeight(44)
        self._clear_btn.setFixedWidth(140)
        self._clear_btn.clicked.connect(self._reset)
        fl.addWidget(self._clear_btn)

        self._open_folder_btn = QtWidgets.QPushButton("Open folder  📂")
        self._open_folder_btn.setObjectName("secondary_btn")
        self._open_folder_btn.setFixedHeight(44)
        self._open_folder_btn.hide()
        self._open_folder_btn.clicked.connect(self._open_output_folder)
        fl.addWidget(self._open_folder_btn)

        self._open_results_btn = QtWidgets.QPushButton("Open results  📄")
        self._open_results_btn.setObjectName("secondary_btn")
        self._open_results_btn.setFixedHeight(44)
        self._open_results_btn.hide()
        self._open_results_btn.clicked.connect(self._open_results_file)
        fl.addWidget(self._open_results_btn)

        self._stop_btn = QtWidgets.QPushButton("Stop")
        self._stop_btn.setObjectName("danger_btn")
        self._stop_btn.setFixedHeight(44)
        self._stop_btn.setFixedWidth(120)
        self._stop_btn.hide()
        self._stop_btn.clicked.connect(self.stopRequested)
        fl.addWidget(self._stop_btn)

        fl.addStretch()

        self._blast_btn = QtWidgets.QPushButton("Run BLAST  →")
        self._blast_btn.setObjectName("primary_btn")
        self._blast_btn.setFixedHeight(44)
        self._blast_btn.setFixedWidth(300)
        self._blast_btn.setEnabled(False)
        self._blast_btn.clicked.connect(self._emit_blast)
        self._blast_btn.setStyleSheet(
            f"QPushButton {{ background-color: {GRAY_LINE}; color: {TEXT_HINT}; border:none; "
            f"border-radius:8px; padding:9px 20px; font-size:18px; font-weight:500; }}"
        )
        fl.addWidget(self._blast_btn)
        outer_layout.addWidget(footer)

        self._last_outdir = ""
        self._last_tsv    = ""

        self.installEventFilter(self)
    
    def eventFilter(self, obj, event):
        """Detectar cuando el panel cambia de tamaño para ajustar el log."""
        if obj == self and event.type() == QtCore.QEvent.Resize:
            self._adjust_log_height()
        return super().eventFilter(obj, event)
    
    def _adjust_log_height(self):
        """Ajustar la altura del log al 30% del panel."""
        if self._log.isVisible():
            target_height = int(self.height() * 0.3)  # 30% del panel
            # Respetar altura mínima
            target_height = max(target_height, 200)
            self._log.setFixedHeight(target_height)
    
    def showEvent(self, event):
        """Cuando el panel se muestra por primera vez."""
        super().showEvent(event)
        self._adjust_log_height()

    # ── API key persistence ───────────────────────────────────────────────

    @staticmethod
    def _config_path():
        return os.path.join(_profiles_dir(), "blast_config.json")

    def _load_api_key(self) -> str:
        try:
            with open(self._config_path(), "r", encoding="utf-8") as f:
                return _json_mod.load(f).get("api_key", "")
        except Exception:
            return ""

    def _save_api_key(self):
        path = self._config_path()
        try:
            data: dict = {}
            if os.path.isfile(path):
                try:
                    with open(path, "r", encoding="utf-8") as f:
                        data = _json_mod.load(f)
                except Exception:
                    pass
            data["api_key"] = self._api_key_edit.text().strip()
            with open(path, "w", encoding="utf-8") as f:
                _json_mod.dump(data, f, indent=2)
        except Exception:
            pass

    def _on_api_key_changed(self, text: str):
        self._lbl_api_warn.setVisible(not bool(text.strip()))

    def _clear_api_key(self):
        self._api_key_edit.clear()
        self._save_api_key()

    # ── UI helpers ────────────────────────────────────────────────────────

    def retranslateUi(self):
        ctx = "BlastPanel"
        self._lbl_title.setText(_tr(ctx, "BLAST NCBI Search"))
        self._lbl_desc.setText(_tr(ctx,
            "BLAST multiFASTA sequences against NCBI. "
            "Drag-and-drop one or more FASTA files (.fa, .fas, .fasta). "
            "Results include top hits with organism and taxonomic classification."))
        self._settings_box.setTitle(_tr(ctx, "BLAST Settings"))
        self._lbl_api.setText(_tr(ctx, "NCBI API Key:"))
        self._lbl_api_warn.setText(_tr(ctx,
            "⚠  A personal NCBI API key is required. Register free at: "
            "ncbi.nlm.nih.gov/account — key is saved automatically."))
        self._lbl_db.setText(_tr(ctx, "Database:"))
        self._lbl_prog.setText(_tr(ctx, "Program:"))
        self._lbl_hits.setText(_tr(ctx, "Hits per sequence (1–100):"))
        self._lbl_batch.setText(_tr(ctx, "Sequences per batch (max 50 recommended):"))
        self._lbl_tax.setText(_tr(ctx, "Taxonomy lookup:"))
        self._tax_check.setText(_tr(ctx, "Fetch organism + taxonomy"))
        self._clear_btn.setText(_tr(ctx, "Clear"))
        self._open_folder_btn.setText(_tr(ctx, "Open folder  📂"))
        self._open_results_btn.setText(_tr(ctx, "Open results  📄"))
        self._blast_btn.setText(_tr(ctx, "Run BLAST  →"))
        self._drop.retranslateUi()

    def changeEvent(self, event):
        if event.type() == QtCore.QEvent.LanguageChange:
            self.retranslateUi()
        super().changeEvent(event)

    # ── Slots ─────────────────────────────────────────────────────────────

    def _on_files(self, paths):
        enabled = len(paths) >= 1
        self._blast_btn.setEnabled(enabled)
        if enabled:
            self._blast_btn.setStyleSheet(
                f"QPushButton {{ background-color: {BLUE}; color: white; border:none; "
                f"border-radius:8px; padding:9px 20px; font-size:18px; font-weight:500; }}"
                f"QPushButton:hover {{ background-color: #0C4A82; }}"
            )
        else:
            self._blast_btn.setStyleSheet(
                f"QPushButton {{ background-color: {GRAY_LINE}; color: {TEXT_HINT}; border:none; "
                f"border-radius:8px; padding:9px 20px; font-size:18px; font-weight:500; }}"
            )

    def _emit_blast(self):
        cfg = {
            "api_key":        self._api_key_edit.text().strip(),
            "database":       self._DATABASES[self._db_combo.currentIndex()],
            "program":        self._PROGRAMS[self._prog_combo.currentIndex()],
            "nhits":          self._hits_spin.value(),
            "nseq":           self._batch_spin.value(),
            "fetch_taxonomy": self._tax_check.isChecked(),
        }
        self.blastRequested.emit(list(self._drop.files), cfg)

    def _open_output_folder(self):
        if self._last_outdir and os.path.isdir(self._last_outdir):
            os.startfile(self._last_outdir)

    def _open_results_file(self):
        if self._last_tsv and os.path.isfile(self._last_tsv):
            os.startfile(self._last_tsv)

    def _reset(self):
        self._drop.clear()
        for k in self._SLOT_KEYS:
            self._log_slots[k] = ""
        self._info_base = ""
        self._elapsed_timer.stop()
        self._log.clear()
        self._log.hide()
        self._open_folder_btn.hide()
        self._open_results_btn.hide()
        self._stop_btn.hide()
        self._blast_btn.show()
        self._blast_btn.setEnabled(False)
        self._blast_btn.setStyleSheet(
            f"QPushButton {{ background-color: {GRAY_LINE}; color: {TEXT_HINT}; border:none; "
            f"border-radius:8px; padding:9px 20px; font-size:18px; font-weight:500; }}"
        )
        self._last_outdir = ""
        self._last_tsv    = ""
        self._clear_btn.setEnabled(True)

    # ── Public API (called by MainWindow) ──────────────────────────────────

    def _rebuild_log(self):
        """Rewrite the entire log widget from the fixed slots dict."""
        sep = "─" * 56
        lines = [
            self._log_slots.get("info",     ""),
            sep,
            self._log_slots.get("blast",    ""),
            self._log_slots.get("organism", ""),
            self._log_slots.get("taxonomy", ""),
            self._log_slots.get("progress", ""),
            sep,
            self._log_slots.get("result",   ""),
        ]
        self._log.setPlainText("\n".join(lines))

    def update_status(self, key: str, text: str):
        """Update a named live slot and refresh the display."""
        self._log.show()
        self._adjust_log_height()       # keep log at ~30% of panel once visible
        if key == "info":
            self._info_base = text      # save static part; timer appends elapsed
        self._log_slots[key] = text
        self._rebuild_log()

    def _tick_elapsed(self):
        """Called every second by the timer to update elapsed time in the info line."""
        elapsed = int(time.monotonic() - self._start_time)
        h, rem  = divmod(elapsed, 3600)
        m, s    = divmod(rem, 60)
        t_str   = f"{h}h {m:02d}:{s:02d}" if h else f"{m:02d}:{s:02d}"
        self._log_slots["info"] = f"{self._info_base}  │  Time: {t_str}"
        self._rebuild_log()

    def set_running(self, running: bool):
        self._blast_btn.setVisible(not running)
        self._stop_btn.setVisible(running)
        self._clear_btn.setEnabled(not running)
        if running:
            self._start_time = time.monotonic()
            self._elapsed_timer.start()
            self._log.show()
            self._adjust_log_height()   # size log to ~30% of panel when run starts
        else:
            self._elapsed_timer.stop()

    def set_progress(self, current: int, total: int):
        if total > 0:
            pct = int(current * 100 / total)
            bar_len = 28
            filled = int(bar_len * current / total)
            bar = "█" * filled + " " * (bar_len - filled)
            self.update_status(
                "progress",
                f"Progress    │ [{bar}] {pct}%"
            )

    def on_finished(self, outdir: str):
        self.set_running(False)
        self._last_outdir = outdir
        if outdir and os.path.isdir(outdir):
            self._open_folder_btn.show()
            # Prefer xlsx; fall back to tsv if xlsx conversion failed
            for ext in (".xlsx", ".tsv"):
                matches = sorted(
                    (os.path.join(outdir, f) for f in os.listdir(outdir)
                     if f.endswith(ext) and f.startswith("blast-")),
                    key=os.path.getmtime, reverse=True
                )
                if matches:
                    self._last_tsv = matches[0]
                    self._open_results_btn.show()
                    break
        # Do NOT overwrite the result slot — the worker already set the final message
        self._blast_btn.setEnabled(bool(self._drop.files))
        if self._drop.files:
            self._blast_btn.setStyleSheet(
                f"QPushButton {{ background-color: {BLUE}; color: white; border:none; "
                f"border-radius:8px; padding:9px 20px; font-size:18px; font-weight:500; }}"
                f"QPushButton:hover {{ background-color: #0C4A82; }}"
            )

    def on_error(self, msg: str):
        self.set_running(False)
        self.update_status("result", f"ERROR       │ {msg[:80]}")
        self._blast_btn.setEnabled(bool(self._drop.files))
        if self._drop.files:
            self._blast_btn.setStyleSheet(
                f"QPushButton {{ background-color: {BLUE}; color: white; border:none; "
                f"border-radius:8px; padding:9px 20px; font-size:18px; font-weight:500; }}"
                f"QPushButton:hover {{ background-color: #0C4A82; }}"
            )


# ═══════════════════════════════════════════════════════════════════════════
# BLAST WORKER
# ═══════════════════════════════════════════════════════════════════════════

class _BlastWorker(QtCore.QThread):
    """
    Python equivalent of blast-tax-remote.sh.
    Submits multiFASTA batches to NCBI BLAST via the CGI API, then retrieves
    organism names (nucleotide DB) and taxonomic classification (taxonomy DB)
    for each hit accession.  Results are saved as a TSV file, written per
    batch so data is never lost if the run is interrupted.
    """
    statusUpdated   = QtCore.pyqtSignal(str, str)   # (slot_key, text)
    progressUpdated = QtCore.pyqtSignal(int, int)   # current, total
    taskFinished    = QtCore.pyqtSignal(str)         # output directory
    taskError       = QtCore.pyqtSignal(str)

    _NCBI_BASE    = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
    _BLAST_URL    = "https://blast.ncbi.nlm.nih.gov/blast/Blast.cgi"
    _MAX_RETRY    = 50
    _POLL_SLEEP   = 5    # seconds between BLAST status polls
    _SOCK_TIMEOUT = 15   # max seconds blocked in a single urlopen call
    _TAX_RETRIES  = 2    # extra retry rounds for "Not_found_in_Taxonomy" results
    _NCBI_RATE    = 9.0  # max HTTP requests/second (NCBI allows 10 with API key)
    _BATCH_UNITS  = 1000 # progress units per batch (phases: wait 0-400, org 400-600, tax 600-800, rows 800-1000)

    def __init__(self, files: List[str], cfg: dict, parent=None):
        super().__init__(parent)
        self.files  = files
        self.cfg    = cfg
        self._stop  = False

        # Shared cache dicts (populated in _run_blast, accessed from threads)
        self._accdb:  dict = {}
        self._taxadb: dict = {}
        self._cache_lock = threading.Lock()
        # Keys already persisted to disk — only new ones are appended per batch
        self._saved_acc_keys:  set = set()
        self._saved_tax_keys:  set = set()
        # Org keys whose CURRENT _taxadb value is a negative that came from a
        # FAILED efetch (transient/network), not from a real "no lineage" answer.
        # These are kept in memory (so the in-run retry loop still finds them) but
        # must NOT be persisted to taxadb.dbx — otherwise a network blip would mark
        # an organism "Not_found_in_Taxonomy" permanently across future runs.
        self._tax_unconfirmed: set = set()

        # Internal map built during batch organism fetch: org_key → taxid string
        # Used to skip the esearch step when fetching taxonomy
        self._org_to_taxid: dict = {}

        # Rate limiter: serialises HTTP calls across all worker threads so
        # we stay within NCBI's 10 req/s limit (using 9 for safety margin).
        self._rl_lock     = threading.Lock()
        self._rl_next     = 0.0   # monotonic time of next allowed request

        # Usage monitor — accessed from multiple threads via _usage_lock
        self._usage_lock       = threading.Lock()
        self._req_count        = 0
        self._rate_limit_count = 0
        self._server_err_count = 0

    def stop(self):
        self._stop = True

    def _rate_acquire(self):
        """Block until the next NCBI request slot is available."""
        interval = 1.0 / self._NCBI_RATE
        with self._rl_lock:
            now  = time.monotonic()
            wait = self._rl_next - now
            if wait > 0:
                time.sleep(wait)
            self._rl_next = time.monotonic() + interval

    # ── HTTP helpers ──────────────────────────────────────────────────────

    def _http_get(self, url, params=None, timeout=60):
        import urllib.request, urllib.parse, urllib.error
        if params:
            url = url + "?" + urllib.parse.urlencode(params)
        for attempt in range(self._MAX_RETRY):
            if self._stop:
                return ""
            self._rate_acquire()
            if self._stop:
                return ""
            try:
                req = urllib.request.Request(url)
                with urllib.request.urlopen(req, timeout=timeout) as resp:
                    if resp.status == 200:
                        text = resp.read().decode("utf-8", errors="replace")
                        if text:
                            with self._usage_lock:
                                self._req_count += 1
                            return text
            except urllib.error.HTTPError as e:
                if e.code == 429:
                    with self._usage_lock:
                        self._rate_limit_count += 1
                    try:
                        wait = int(e.headers.get("Retry-After") or 0)
                    except Exception:
                        wait = 0
                    if wait <= 0:
                        wait = min(30 * (attempt + 1), 120)
                    self._interruptible_sleep(wait)
                elif 500 <= e.code < 600:
                    with self._usage_lock:
                        self._server_err_count += 1
                    self._interruptible_sleep(min(10 * (attempt + 1), 60))
                else:
                    return ""  # 4xx other than 429 — not retryable
            except Exception:
                self._interruptible_sleep(2)
        return ""

    def _http_post(self, url, data: str, timeout=120):
        import urllib.request, urllib.error
        for attempt in range(self._MAX_RETRY):
            if self._stop:
                return ""
            self._rate_acquire()
            if self._stop:
                return ""
            try:
                req = urllib.request.Request(
                    url,
                    data=data.encode("utf-8"),
                    headers={"Content-Type": "application/x-www-form-urlencoded"},
                    method="POST"
                )
                with urllib.request.urlopen(req, timeout=timeout) as resp:
                    if resp.status == 200:
                        text = resp.read().decode("utf-8", errors="replace")
                        if text:
                            with self._usage_lock:
                                self._req_count += 1
                            return text
            except urllib.error.HTTPError as e:
                if e.code == 429:
                    with self._usage_lock:
                        self._rate_limit_count += 1
                    try:
                        wait = int(e.headers.get("Retry-After") or 0)
                    except Exception:
                        wait = 0
                    if wait <= 0:
                        wait = min(30 * (attempt + 1), 120)
                    self._interruptible_sleep(wait)
                elif 500 <= e.code < 600:
                    with self._usage_lock:
                        self._server_err_count += 1
                    self._interruptible_sleep(min(10 * (attempt + 1), 60))
                else:
                    return ""  # 4xx other than 429 — not retryable
            except Exception:
                self._interruptible_sleep(2)
        return ""

    def _interruptible_sleep(self, seconds: float):
        """Sleep in 0.5 s chunks so _stop is checked frequently."""
        end = time.monotonic() + seconds
        while time.monotonic() < end:
            if self._stop:
                return
            time.sleep(0.5)

    # ── FASTA helpers ─────────────────────────────────────────────────────

    def _merge_fasta_files(self, files):
        parts = []
        for f in files:
            try:
                with open(f, "r", encoding="utf-8", errors="replace") as fh:
                    content = fh.read().replace("\r\n", "\n").replace("\r", "\n")
                    parts.append(content.strip())
            except Exception as e:
                self.statusUpdated.emit("blast", f"BLAST       │ Warning: could not read {os.path.basename(f)}")
        return "\n".join(parts)

    def _to_single_line_fasta(self, text):
        out = []
        header = None
        seq_parts = []
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if header is not None:
                    out.append(header)
                    out.append("".join(seq_parts))
                header = line.replace(" ", "_")
                seq_parts = []
            else:
                seq_parts.append(line)
        if header is not None:
            out.append(header)
            out.append("".join(seq_parts))
        return "\n".join(out)

    def _split_batches(self, fasta_text, nseq):
        """Return (batches, batch_pairs_list).
        batches[i]       – FASTA text string for batch i
        batch_pairs_list[i] – list of (header, seq) tuples for batch i
        """
        # `fasta_text` comes from _to_single_line_fasta(): each header is followed by
        # exactly one sequence line (possibly empty). Do NOT drop blank lines here —
        # filtering an empty sequence line would shift the next header into its place
        # and desync every header/seq pair from that point on.
        lines = fasta_text.splitlines()
        pairs = []
        i = 0
        while i < len(lines):
            if lines[i].startswith(">"):
                header = lines[i]
                seq    = lines[i + 1] if i + 1 < len(lines) else ""
                pairs.append((header, seq))
                i += 2
            else:
                i += 1
        batches = []
        batch_pairs_list = []
        for start in range(0, len(pairs), nseq):
            chunk = pairs[start:start + nseq]
            batches.append("\n".join(h + "\n" + s for h, s in chunk))
            batch_pairs_list.append(chunk)
        return batches, batch_pairs_list

    # ── BLAST API ─────────────────────────────────────────────────────────

    def _url_encode_fasta(self, text):
        from urllib.parse import quote
        safe = (
            "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
            "abcdefghijklmnopqrstuvwxyz"
            "0123456789_-"
        )
        return quote(text, safe=safe)

    def _blast_submit(self, fasta_text):
        cfg = self.cfg
        encoded = self._url_encode_fasta(fasta_text)
        data = (
            f"CMD=Put&PROGRAM={cfg['program']}&DATABASE={cfg['database']}"
            f"&api_key={cfg['api_key']}&HITLIST_SIZE={cfg['nhits']}&QUERY={encoded}"
        )
        resp = self._http_post(self._BLAST_URL, data)
        if not resp:
            return None, 30
        import re
        rid_m  = re.search(r"RID = ([^\s]+)\s+RTOE", resp)
        rtoe_m = re.search(r"RTOE = (\d+)", resp)
        rid  = rid_m.group(1) if rid_m else None
        rtoe = int(rtoe_m.group(1)) if rtoe_m else 30
        return rid, rtoe

    def _blast_poll(self, rid, batch_label=""):
        url = f"{self._BLAST_URL}?CMD=Get&RID={rid}"
        attempts = 0
        t0 = time.monotonic()
        prefix = f"BLAST       │ [{batch_label}] " if batch_label else "BLAST       │ "
        while not self._stop and attempts < self._MAX_RETRY:
            resp = self._http_get(url)
            if self._stop:
                return False
            if "Status=WAITING" in resp:
                elapsed = int(time.monotonic() - t0)
                self.statusUpdated.emit(
                    "blast",
                    f"{prefix}Polling for results… "
                    f"(attempt {attempts + 1}/{self._MAX_RETRY} · {elapsed}s elapsed)"
                )
                self._interruptible_sleep(self._POLL_SLEEP)
                attempts += 1
                continue
            if "Status=FAILED" in resp:
                self.statusUpdated.emit("blast", f"{prefix}Search failed for RID {rid}.")
                return False
            if "Status=UNKNOWN" in resp:
                self.statusUpdated.emit("blast", f"{prefix}Search expired for RID {rid}.")
                return False
            if "Status=READY" in resp:
                return True
            # Empty or unrecognised response — count as a transient failure
            elapsed = int(time.monotonic() - t0)
            self.statusUpdated.emit(
                "blast",
                f"{prefix}Polling for results… "
                f"(attempt {attempts + 1}/{self._MAX_RETRY} · {elapsed}s elapsed)"
            )
            attempts += 1
            self._interruptible_sleep(self._POLL_SLEEP)
        if not self._stop:
            elapsed = int(time.monotonic() - t0)
            self.statusUpdated.emit(
                "blast",
                f"{prefix}No response after {attempts} polls ({elapsed}s)."
            )
        return False

    def _blast_get_tabular(self, rid):
        url = (
            f"{self._BLAST_URL}"
            f"?CMD=Get&FORMAT_TYPE=Text&ALIGNMENT_VIEW=Tabular&RID={rid}"
        )
        resp = self._http_get(url, timeout=120)
        return self._parse_tabular(resp)

    def _parse_tabular(self, text):
        # Equivalent to: grep -A nhits '^#' | sed '/^#/d; /^--/d'
        rows = []
        nhits   = self.cfg["nhits"]
        capture = False
        count   = 0
        for line in text.splitlines():
            if line.startswith("# Query:") or (
                line.startswith("#") and "Fields:" in line
            ):
                capture = True
                count   = 0
                continue
            if capture:
                if line.startswith("#") or line.startswith("--"):
                    continue
                stripped = line.strip()
                if stripped:
                    rows.append(stripped)
                    count += 1
                    if count >= nhits:
                        capture = False
        return rows

    @staticmethod
    def _rank_rows(rows):
        """Prepend a per-sample Hit_rank column (1 = best hit) to each tabular row.

        `rows` must arrive in BLAST quality order (best first) within each query,
        as produced by _parse_tabular — so the rank is assigned BEFORE any
        alphabetical reordering. The output is grouped by Query_name with hits in
        ascending rank, and each row gains the rank as its first tab field."""
        counter = {}
        ranked = []
        for row in rows:
            query = row.split("\t", 1)[0]
            rank = counter.get(query, 0) + 1
            counter[query] = rank
            ranked.append((query, rank, row))
        ranked.sort(key=lambda t: (t[0], t[1]))
        return [f"{rank}\t{row}" for (query, rank, row) in ranked]

    # ── NCBI metadata ────────────────────────────────────────────────────

    def _fetch_organism(self, accession: str) -> str:
        """Return organism name for *accession*, using in-memory cache."""
        with self._cache_lock:
            if accession in self._accdb:
                return self._accdb[accession]
        if self._stop:
            return ""
        import xml.etree.ElementTree as ET
        base = self._NCBI_BASE
        xml = self._http_get(
            f"{base}esearch.fcgi",
            params={"db": "nucleotide", "term": accession, "usehistory": "y"}
        )
        if not xml or self._stop:
            return ""   # transient failure — do NOT cache so a later batch can retry
        try:
            root = ET.fromstring(xml)
            web = root.findtext("WebEnv", "")
            key = root.findtext("QueryKey", "")
        except Exception:
            return ""
        if not web or not key:
            return ""
        xml2 = self._http_get(
            f"{base}efetch.fcgi",
            params={
                "db": "nucleotide", "query_key": key,
                "WebEnv": web, "rettype": "gbc", "retmode": "xml"
            },
            timeout=90
        )
        if not xml2 or self._stop:
            return ""   # transient failure — do NOT cache
        try:
            root2 = ET.fromstring(xml2)
            el = root2.find(".//INSDSeq_organism")
            if el is None or not el.text:
                return ""
            org = (
                el.text.strip()
                .replace(" ", "_")
                .replace("&apos;", "'")
                .replace("&amp;", "&")
            )
        except Exception:
            return ""
        with self._cache_lock:
            self._accdb[accession] = org
        return org

    def _fetch_taxonomy(self, organism: str) -> str:
        """Return taxonomy string for *organism*, using in-memory cache."""
        with self._cache_lock:
            if organism in self._taxadb:
                return self._taxadb[organism]
        if self._stop:
            return "Not_found_in_Taxonomy"
        import xml.etree.ElementTree as ET
        base = self._NCBI_BASE
        # NCBI scientific names use spaces; organism is stored with underscores
        search_name = organism.replace("_", " ")
        xml = self._http_get(
            f"{base}esearch.fcgi",
            params={
                "db": "taxonomy",
                "term": f'"{search_name}"[Scientific Name]',
                "usehistory": "y"
            }
        )
        if not xml or self._stop:
            return "Not_found_in_Taxonomy"   # transient — do NOT cache
        try:
            root = ET.fromstring(xml)
            web = root.findtext("WebEnv", "")
            key = root.findtext("QueryKey", "")
        except Exception:
            return "Not_found_in_Taxonomy"
        if not web or not key:
            return "Not_found_in_Taxonomy"
        xml2 = self._http_get(
            f"{base}efetch.fcgi",
            params={
                "db": "taxonomy", "query_key": key,
                "WebEnv": web, "retmode": "xml"
            },
            timeout=90
        )
        if not xml2 or self._stop:
            return "Not_found_in_Taxonomy"   # transient — do NOT cache
        tax = self._parse_taxonomy_xml(xml2)
        result = tax if tax else "Not_found_in_Taxonomy"
        # Cache confirmed results (both found and genuinely absent)
        with self._cache_lock:
            self._taxadb[organism] = result
        return result

    def _parse_taxonomy_xml(self, xml_text):
        import xml.etree.ElementTree as ET
        try:
            root = ET.fromstring(xml_text)
        except Exception:
            return ""
        ranks_wanted = {"kingdom", "class", "order", "family", "genus"}
        found = {}
        for taxon in root.iter("Taxon"):
            rank = (taxon.findtext("Rank") or "").strip().lower()
            name = (taxon.findtext("ScientificName") or "").strip()
            if rank in ranks_wanted and rank not in found:
                found[rank] = name
        return "\t".join(
            found.get(r, "-") for r in ["kingdom", "class", "order", "family", "genus"]
        )

    # ── Batch fetch methods ───────────────────────────────────────────────

    def _fetch_organisms_batch(self, accessions: List[str]) -> None:
        """Fetch organism names + taxids for all accessions in one POST per 500-ID chunk.

        Populates _accdb (accession → org_key) and _org_to_taxid (org_key → taxid).
        Accessions already in _accdb are skipped.
        """
        import xml.etree.ElementTree as ET
        import urllib.parse as _ulp

        to_fetch = [a for a in accessions if a not in self._accdb]
        if not to_fetch:
            return

        api_key = self.cfg.get("api_key", "").strip()
        base    = self._NCBI_BASE

        for start in range(0, len(to_fetch), 500):
            if self._stop:
                return
            chunk  = to_fetch[start:start + 500]
            params: dict = {"db": "nucleotide", "id": ",".join(chunk),
                            "rettype": "gbc", "retmode": "xml"}
            if api_key:
                params["api_key"] = api_key
            xml_text = self._http_post(f"{base}efetch.fcgi",
                                       _ulp.urlencode(params), timeout=120)
            if not xml_text or self._stop:
                continue
            try:
                root = ET.fromstring(xml_text)
            except Exception:
                continue
            with self._cache_lock:
                for seq in root.iter("INSDSeq"):
                    accver   = (seq.findtext("INSDSeq_accession-version") or "").strip()
                    organism = (seq.findtext("INSDSeq_organism") or "").strip()
                    if not accver or not organism:
                        continue
                    org_key = (organism.replace(" ", "_")
                                       .replace("&apos;", "'")
                                       .replace("&amp;", "&"))
                    self._accdb[accver] = org_key
                    # Extract taxid from the source feature db_xref qualifier
                    for feat in seq.iter("INSDFeature"):
                        if feat.findtext("INSDFeature_key", "") == "source":
                            for qual in feat.iter("INSDQualifier"):
                                if qual.findtext("INSDQualifier_name", "") == "db_xref":
                                    val = qual.findtext("INSDQualifier_value", "")
                                    if val.startswith("taxon:"):
                                        self._org_to_taxid[org_key] = val.split(":")[1].strip()
                                        break
                            break

    def _fetch_taxonomy_batch(self, organisms: List[str]) -> None:
        """Fetch taxonomy for all organisms using taxids when available.

        Fast path: organisms in _org_to_taxid → single batch efetch on taxonomy DB.
        Fallback: organisms without a known taxid use the existing esearch-based
        _fetch_taxonomy() (rare — only for accessions loaded from .dbx cache files
        created by a previous run before this batch optimisation was added).
        Populates _taxadb (org_key → tab-separated ranks).
        """
        import urllib.parse as _ulp

        to_fetch = [o for o in organisms if o not in self._taxadb]
        if not to_fetch:
            return

        api_key = self.cfg.get("api_key", "").strip()
        base    = self._NCBI_BASE

        with_taxid    = [(o, self._org_to_taxid[o]) for o in to_fetch
                         if o in self._org_to_taxid]
        without_taxid = [o for o in to_fetch if o not in self._org_to_taxid]

        # ── Fast path: batch efetch by taxid ─────────────────────────────
        unique_taxids = list(dict.fromkeys(tid for _, tid in with_taxid))
        taxid_results: Dict[str, str] = {}
        # Taxids whose efetch chunk did NOT return a usable response (network
        # failure, stop, etc.). A negative for these is transient, not confirmed.
        failed_taxids: set = set()
        for start in range(0, len(unique_taxids), 500):
            if self._stop:
                # Treat the rest as transient so they are retried, never persisted.
                failed_taxids.update(unique_taxids[start:])
                break
            chunk  = unique_taxids[start:start + 500]
            params: dict = {"db": "taxonomy", "id": ",".join(chunk), "retmode": "xml"}
            if api_key:
                params["api_key"] = api_key
            xml_text = self._http_post(f"{base}efetch.fcgi",
                                       _ulp.urlencode(params), timeout=120)
            if xml_text and not self._stop:
                taxid_results.update(self._parse_taxonomy_xml_batch(xml_text))
            else:
                failed_taxids.update(chunk)

        with self._cache_lock:
            for org, taxid in with_taxid:
                if taxid in taxid_results:
                    # Confirmed answer (lineage found) from a successful response.
                    self._taxadb[org] = taxid_results[taxid]
                    self._tax_unconfirmed.discard(org)
                elif taxid in failed_taxids:
                    # Transient failure: keep a negative in memory so the in-run
                    # retry loop still finds it, but mark it unconfirmed so it is
                    # excluded from persistence to taxadb.dbx.
                    self._taxadb[org] = "Not_found_in_Taxonomy"
                    self._tax_unconfirmed.add(org)
                else:
                    # Chunk succeeded but this taxid was absent from the response =
                    # genuine "no lineage" negative → safe to cache and persist.
                    self._taxadb[org] = "Not_found_in_Taxonomy"
                    self._tax_unconfirmed.discard(org)

        # ── Fallback: esearch by scientific name (legacy cache hits) ──────
        for org in without_taxid:
            if self._stop:
                return
            self._fetch_taxonomy(org)

    def _parse_taxonomy_xml_batch(self, xml_text: str) -> Dict[str, str]:
        """Parse a multi-taxon taxonomy XML; returns {taxid_str: tab-separated ranks}."""
        import xml.etree.ElementTree as ET
        try:
            root = ET.fromstring(xml_text)
        except Exception:
            return {}
        ranks_wanted = {"kingdom", "class", "order", "family", "genus"}
        results: Dict[str, str] = {}
        for taxon_el in root.findall("Taxon"):   # direct children = one per requested taxid
            taxid = (taxon_el.findtext("TaxId") or "").strip()
            if not taxid:
                continue
            found: Dict[str, str] = {}
            for child in taxon_el.iter("Taxon"):  # includes LineageEx descendants
                rank = (child.findtext("Rank") or "").strip().lower()
                name = (child.findtext("ScientificName") or "").strip()
                if rank in ranks_wanted and rank not in found:
                    found[rank] = name
            lineage = "\t".join(
                found.get(r, "-") for r in ["kingdom", "class", "order", "family", "genus"]
            )
            results[taxid] = lineage
            # NCBI returns a merged taxon under its CURRENT TaxId; a requested taxid
            # that was merged is listed under <AkaTaxIds>. Map those too so the
            # caller's lookup by the requested id resolves instead of being wrongly
            # cached/persisted as "Not_found_in_Taxonomy".
            aka = taxon_el.find("AkaTaxIds")
            if aka is not None:
                for alt_el in aka.findall("TaxId"):
                    alt = (alt_el.text or "").strip()
                    if alt:
                        results[alt] = lineage
        return results

    # ── Cache I/O ────────────────────────────────────────────────────────

    def _load_cache(self, path):
        cache = {}
        if os.path.isfile(path):
            try:
                with open(path, "r", encoding="utf-8") as fh:
                    for line in fh:
                        parts = line.rstrip("\n").split("\t", 1)
                        if len(parts) == 2:
                            cache[parts[0]] = parts[1]
            except Exception:
                pass
        return cache

    def _append_cache(self, new_entries: dict, path: str):
        """Append only new key-value pairs to the cache file (no full rewrite)."""
        if not new_entries:
            return
        try:
            with open(path, "a", encoding="utf-8") as fh:
                for k in sorted(new_entries):
                    fh.write(f"{k}\t{new_entries[k]}\n")
        except Exception:
            pass

    # ── TSV → XLSX conversion ─────────────────────────────────────────────

    def _tsv_to_xlsx(self, tsv_path: str) -> str:
        """Convert *tsv_path* to a formatted xlsx. Returns xlsx path or '' on failure."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            self.statusUpdated.emit("result", f"XLSX skip  │ openpyxl not available: {e}")
            return ""
        try:
            with open(tsv_path, "r", encoding="utf-8") as fh:
                lines = [l for l in fh.read().splitlines() if l.strip()]
        except Exception as e:
            self.statusUpdated.emit("result", f"XLSX skip  │ could not read TSV: {e}")
            return ""
        if not lines:
            return ""

        # The TSV already carries Hit_rank (per-sample 1..N, best hit = 1) as its
        # first column, so the layout is read as-is.
        headers = lines[0].split("\t")
        n_cols  = len(headers)

        # openpyxl 3.x requires 8-char ARGB hex strings (alpha + RGB).
        # Zone 1 → cols 0-1   : Hit_rank, Query_name
        # Zone 2 → cols 2-12  : BLAST metric columns (accession … bit-score)
        # Zone 3 → cols 13+   : Taxonomy columns (only when fetch_taxonomy=True)
        H1 = PatternFill(patternType="solid", fgColor="FF1A365D")   # header navy
        H2 = PatternFill(patternType="solid", fgColor="FF0D5E6E")   # header teal
        H3 = PatternFill(patternType="solid", fgColor="FF7C3200")   # header burnt-orange
        D1 = PatternFill(patternType="solid", fgColor="FFE8F1FB")   # data light-blue
        D2 = PatternFill(patternType="solid", fgColor="FFE8F5F6")   # data light-teal
        D3 = PatternFill(patternType="solid", fgColor="FFFEF3E8")   # data light-orange

        white_bold  = Font(color="FFFFFFFF", bold=True, size=10)
        normal_font = Font(size=10)
        bold_font   = Font(size=10, bold=True)   # best hit (Hit_rank == 1) per sample
        hdr_align   = Alignment(horizontal="center", vertical="center")
        dat_align   = Alignment(vertical="center", wrap_text=False)
        thin        = Side(style="thin", color="FFCCCCCC")
        medium      = Side(style="medium", color="FF9AA0A6")   # sample-block divider
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Same as `border` but with a heavier top edge — marks the first row of each
        # new sample block so the hit groups are visually separated.
        border_group_top = Border(left=thin, right=thin, top=medium, bottom=thin)

        # Hit_rank and the BLAST metric columns (P_identity … Bit_score) are
        # numeric. Writing them as strings makes Excel flag every cell with
        # "Number stored as text", whose background error-checker re-scans the
        # sheet on every sort/filter/scroll → high CPU. Convert these to
        # int/float so openpyxl writes native numeric cells; the rest stay text.
        # Detect them by header name so it is robust to column shifts.
        _NUMERIC_NAMES = frozenset({
            "Hit_rank", "P_identity", "Alignment_length", "Num_mismatches",
            "Gap_opens", "Query_start", "Query_end", "Subject_start",
            "Subject_end", "Evalue", "Bit_score",
        })
        _numeric_idx = frozenset(
            i for i, h in enumerate(headers) if h in _NUMERIC_NAMES)

        def _num(v):
            if v == "":
                return v
            try:
                return int(v)
            except ValueError:
                pass
            try:
                return float(v)   # handles decimals and e-notation (Evalue)
            except ValueError:
                return v          # leave genuinely non-numeric text as-is

        def _hfill(ci):
            return H1 if ci <= 1 else (H2 if ci <= 12 else H3)

        def _dfill(ci, tinted):
            if not tinted:
                return None
            return D1 if ci <= 1 else (D2 if ci <= 12 else D3)

        wb = Workbook()
        ws = wb.active
        ws.title = "BLAST Results"

        # Header row
        ws.append(headers)
        for ci in range(n_cols):
            cell = ws.cell(row=1, column=ci + 1)
            cell.fill      = _hfill(ci)
            cell.font      = white_bold
            cell.alignment = hdr_align
            cell.border    = border
        ws.row_dimensions[1].height = 22

        # Data rows — shaded in BLOCKS by sample (Query_name), not per single row,
        # so each sample's hits share one tint and the next sample flips tint. A
        # heavier top border marks the first row of each new block. Rows already
        # arrive grouped by Query_name (see _rank_rows), so a simple value-change
        # check delimits the blocks. Column indices are looked up by name to stay
        # robust to layout shifts (taxonomy columns present or not).
        try:
            _query_idx = headers.index("Query_name")
        except ValueError:
            _query_idx = 1
        try:
            _rank_idx = headers.index("Hit_rank")
        except ValueError:
            _rank_idx = 0

        prev_query   = None
        block_tinted = True   # first block tinted (matches former row-2 behaviour)
        for rn, line in enumerate(lines[1:], start=2):
            raw_vals = line.split("\t")
            while len(raw_vals) < n_cols:
                raw_vals.append("")
            raw_vals = raw_vals[:n_cols]

            query     = raw_vals[_query_idx] if _query_idx < n_cols else ""
            new_block = (prev_query is not None and query != prev_query)
            if new_block:
                block_tinted = not block_tinted
            prev_query = query

            rank_raw = raw_vals[_rank_idx] if _rank_idx < n_cols else ""
            is_best  = (str(rank_raw).strip() == "1")   # best hit of this sample

            vals = [
                _num(v) if ci in _numeric_idx else v
                for ci, v in enumerate(raw_vals)
            ]
            ws.append(vals)
            row_border = border_group_top if new_block else border
            row_font   = bold_font if is_best else normal_font
            for ci in range(n_cols):
                cell = ws.cell(row=rn, column=ci + 1)
                fill = _dfill(ci, block_tinted)
                if fill:
                    cell.fill = fill
                cell.font      = row_font
                cell.alignment = dat_align
                cell.border    = row_border

        # Freeze header, enable auto-filter
        ws.freeze_panes    = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Auto-fit column widths (capped at 55 chars)
        for ci, col_cells in enumerate(ws.columns):
            width = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(ci + 1)].width = min(width + 2, 55)

        xlsx_path = tsv_path.rsplit(".", 1)[0] + ".xlsx"
        try:
            wb.save(xlsx_path)
        except Exception as e:
            self.statusUpdated.emit("result", f"XLSX error │ {e}")
            return ""
        return xlsx_path

    # ── Main run ─────────────────────────────────────────────────────────

    def run(self):
        try:
            self._run_blast()
        except Exception as e:
            import traceback
            self.taskError.emit(f"{e}\n{traceback.format_exc()}")

    def _run_blast(self):
        cfg       = self.cfg
        nhits     = cfg["nhits"]
        nseq      = cfg["nseq"]
        run_start = datetime.datetime.now()
        mydate    = run_start.strftime("%Y%m%d-%H%M%S")

        # ── Output directory (passed from MainWindow dialog) ──
        output_dir = cfg["outdir"]
        os.makedirs(output_dir, exist_ok=True)

        # ── Cache files stored alongside the output folder ──
        taxadb_path = os.path.join(output_dir, "taxadb.dbx")
        accdb_path  = os.path.join(output_dir, "accdb.dbx")
        fetch_tax = cfg.get("fetch_taxonomy", True)
        if fetch_tax:
            # Pre-load any .dbx files from sibling blast result folders
            parent = os.path.dirname(output_dir)
            for root_dir, _dirs, fnames in os.walk(parent):
                if "taxadb.dbx" in fnames:
                    self._taxadb.update(
                        self._load_cache(os.path.join(root_dir, "taxadb.dbx"))
                    )
                if "accdb.dbx" in fnames:
                    self._accdb.update(
                        self._load_cache(os.path.join(root_dir, "accdb.dbx"))
                    )
            # Snapshot keys already on disk — only new ones will be appended
            self._saved_tax_keys = set(self._taxadb.keys())
            self._saved_acc_keys = set(self._accdb.keys())

        # ── Merge & normalize FASTA ──
        raw   = self._merge_fasta_files(self.files)
        fasta = self._to_single_line_fasta(raw)
        seq_count = fasta.count("\n>") + (1 if fasta.startswith(">") else 0)

        if seq_count == 0:
            self.taskError.emit(
                "No FASTA sequences found in the provided files."
            )
            return

        # ── Split into batches ──
        batches, batch_pairs_list = self._split_batches(fasta, nseq)
        n_batches = len(batches)
        completed_batches: set = set()   # indices of batches fully written to TSV

        # ── Summary line (fixed slot "info") ──
        self.statusUpdated.emit(
            "info",
            f"Sequences: {seq_count}  │  Batches: {n_batches}"
            f"  │  Hits/seq: {nhits}  │  DB: {cfg['database']}"
        )

        _blast_cols = (
            "Query_name\tSubject_accession.ver\tP_identity\tAlignment_length\t"
            "Num_mismatches\tGap_opens\tQuery_start\tQuery_end\t"
            "Subject_start\tSubject_end\tEvalue\tBit_score"
        )
        headings = (
            _blast_cols + "\tSubject_Kingdom\tSubject_Class\tSubject_Order\t"
            "Subject_Family\tSubject_Genus\tSubject_organism"
            if fetch_tax else _blast_cols
        )
        # Per-sample hit rank (1 = best hit) as the first column, so results can
        # be filtered by rank (e.g. Hit_rank == 1 keeps only each sample's top hit).
        headings = "Hit_rank\t" + headings

        # ── Open TSV for incremental writing ──
        tsv_path = os.path.join(output_dir, f"blast-{mydate}.tsv")
        for _attempt in range(10):
            try:
                with open(tsv_path, "w", encoding="utf-8") as tsv_fh:
                    tsv_fh.write(headings + "\n")
                break
            except PermissionError:
                self.statusUpdated.emit(
                    "result",
                    f"⚠ Output file locked — close it to continue… ({_attempt + 1}/10)"
                )
                self._interruptible_sleep(0.5)
        else:
            # Sin la cabecera no tiene sentido lanzar las consultas BLAST (red):
            # se produciría un TSV sin encabezado. Abortar limpiamente.
            self.taskError.emit(
                f"Could not write output file (locked/permission denied):\n{tsv_path}"
            )
            return

        total_hits_done: int = 0
        total_expected = n_batches * self._BATCH_UNITS

        for batch_idx, batch_fasta in enumerate(batches):
            if self._stop:
                break

            batch_seq   = batch_fasta.count(">")
            batch_label = f"Batch {batch_idx+1}/{n_batches}"
            batch_base  = batch_idx * self._BATCH_UNITS

            # ── BLAST ──
            self.statusUpdated.emit(
                "blast",
                f"BLAST       │ [{batch_label}] Submitting {batch_seq} sequences…"
            )
            rid, rtoe = self._blast_submit(batch_fasta)
            if not rid:
                self.statusUpdated.emit(
                    "blast",
                    f"BLAST       │ [{batch_label}] Submission failed — skipping."
                )
                continue

            self.statusUpdated.emit(
                "blast",
                f"BLAST       │ [{batch_label}] RID={rid}  waiting {rtoe}s…"
            )
            for tick in range(rtoe):
                if self._stop:
                    break
                time.sleep(1)
                self.progressUpdated.emit(
                    batch_base + int(400 * (tick + 1) / max(rtoe, 1)),
                    total_expected
                )
            if self._stop:
                break

            self.progressUpdated.emit(batch_base + 400, total_expected)
            if not self._blast_poll(rid, batch_label):
                continue

            blast_rows = self._blast_get_tabular(rid)
            self.statusUpdated.emit(
                "blast",
                f"BLAST       │ [{batch_label}] {len(blast_rows)} hits retrieved  ✓"
            )

            if fetch_tax:
                # ── Fetch organisms (batch: one POST for all accessions) ──
                accessions = [
                    (row.split("\t")[1] if "\t" in row else "") for row in blast_rows
                ]
                unique_accs   = list(dict.fromkeys(a for a in accessions if a))
                n_unique_accs = len(unique_accs)
                self.statusUpdated.emit(
                    "organism",
                    f"Organism ID │ [{batch_label}] Fetching {n_unique_accs} accessions…"
                )
                self._fetch_organisms_batch(unique_accs)
                if self._stop:
                    break
                n_org_found = sum(1 for a in unique_accs if a in self._accdb)
                self.statusUpdated.emit(
                    "organism",
                    f"Organism ID │ [{batch_label}] {n_org_found}/{n_unique_accs} resolved  ✓"
                )
                self.progressUpdated.emit(batch_base + 600, total_expected)

                organisms: List[str] = [
                    (self._accdb.get(acc, "") if acc else "") for acc in accessions
                ]

                # ── Fetch taxonomy (batch: one POST per 500 taxids) ───────
                unique_orgs   = list(dict.fromkeys(o for o in organisms if o))
                n_unique_orgs = len(unique_orgs)
                self.statusUpdated.emit(
                    "taxonomy",
                    f"Taxonomy    │ [{batch_label}] Fetching {n_unique_orgs} organisms…"
                )
                self._fetch_taxonomy_batch(unique_orgs)
                if self._stop:
                    break

                # ── Retry not-found organisms (batch) ─────────────────────
                for attempt in range(1, self._TAX_RETRIES + 1):
                    if self._stop:
                        break
                    retry_orgs = [
                        org for org in unique_orgs
                        if self._taxadb.get(org) == "Not_found_in_Taxonomy"
                    ]
                    if not retry_orgs:
                        break
                    n_retry = len(retry_orgs)
                    with self._cache_lock:
                        for org in retry_orgs:
                            del self._taxadb[org]
                    self.statusUpdated.emit(
                        "taxonomy",
                        f"Taxonomy    │ [{batch_label}] Retry {attempt}/{self._TAX_RETRIES}: {n_retry} not-found…"
                    )
                    self._fetch_taxonomy_batch(retry_orgs)

                n_tax_found = sum(
                    1 for o in unique_orgs
                    if self._taxadb.get(o, "Not_found_in_Taxonomy") != "Not_found_in_Taxonomy"
                )
                self.statusUpdated.emit(
                    "taxonomy",
                    f"Taxonomy    │ [{batch_label}] {n_tax_found}/{n_unique_orgs} resolved  ✓"
                )
                self.progressUpdated.emit(batch_base + 800, total_expected)

                taxonomies: List[str] = [
                    (self._taxadb.get(o, "Not_found_in_Taxonomy") if o else "Not_found_in_Taxonomy")
                    for o in organisms
                ]

                batch_rows_out = self._rank_rows([
                    f"{row}\t{tax}\t{org}"
                    for row, tax, org in zip(blast_rows, taxonomies, organisms)
                ])

                # ── Persist only new entries after every batch ──
                # Exclude _tax_unconfirmed: negatives from a failed efetch are kept
                # in memory for the in-run retry but must not poison taxadb.dbx.
                with self._cache_lock:
                    new_tax = {k: self._taxadb[k] for k in self._taxadb
                               if k not in self._saved_tax_keys
                               and k not in self._tax_unconfirmed}
                    new_acc = {k: self._accdb[k]  for k in self._accdb  if k not in self._saved_acc_keys}
                    self._append_cache(new_tax, taxadb_path)
                    self._append_cache(new_acc,  accdb_path)
                    self._saved_tax_keys.update(new_tax.keys())
                    self._saved_acc_keys.update(new_acc.keys())

            else:
                batch_rows_out = self._rank_rows(blast_rows)

            # ── Append batch rows to disk immediately ──
            row_phase_start = 800 if fetch_tax else 400
            row_phase_range = 200 if fetch_tax else 600
            n_batch_rows = max(len(batch_rows_out), 1)
            _batch_written = False
            for _attempt in range(20):   # retry up to 10 s if TSV is open in Excel
                try:
                    with open(tsv_path, "a", encoding="utf-8") as tsv_fh:
                        for row_idx, r in enumerate(batch_rows_out, 1):
                            tsv_fh.write(r + "\n")
                            self.progressUpdated.emit(
                                batch_base + row_phase_start + int(row_phase_range * row_idx / n_batch_rows),
                                total_expected
                            )
                    _batch_written = True
                    break  # write succeeded
                except PermissionError:
                    self.statusUpdated.emit(
                        "result",
                        f"⚠ TSV file is open — close it and the run will resume… ({_attempt + 1}/20)"
                    )
                    self._interruptible_sleep(0.5)

            # Solo marcar el batch como completado si sus filas llegaron al disco;
            # de lo contrario sus secuencias deben aparecer en el FASTA de "missing".
            # Contar los hits una sola vez, tras la escritura exitosa, para que un
            # reintento (que reescribe el lote completo) no infle el total.
            if _batch_written:
                completed_batches.add(batch_idx)
                total_hits_done += len(batch_rows_out)

        # ── Build missing-sequences FASTA (unprocessed or failed batches) ──
        missing_pairs = []
        for bi, bp in enumerate(batch_pairs_list):
            if bi not in completed_batches:
                missing_pairs.extend(bp)

        miss_msg = ""
        if missing_pairs:
            miss_path = os.path.join(output_dir, f"missing_seqs_{mydate}.fa")
            try:
                with open(miss_path, "w", encoding="utf-8") as fh:
                    for h, s in missing_pairs:
                        fh.write(h + "\n" + s + "\n")
                miss_msg = (
                    f"{len(missing_pairs)} unprocessed seqs → "
                    f"{os.path.basename(miss_path)}"
                )
            except Exception as exc:
                miss_msg = f"Could not write missing FASTA: {exc}"

        # ── Convert TSV → XLSX ──
        xlsx_path = ""
        if not self._stop and total_hits_done > 0:
            xlsx_path = self._tsv_to_xlsx(tsv_path)

        if self._stop:
            result_msg = f"Stopped     │ {miss_msg}" if miss_msg else "Stopped by user."
        else:
            out_name = os.path.basename(xlsx_path if xlsx_path else tsv_path)
            if miss_msg:
                result_msg = f"Done  ✓     │ {total_hits_done} hits written  │ {miss_msg}"
            else:
                result_msg = f"Done  ✓     │ {total_hits_done} hits written → {out_name}"
        self.statusUpdated.emit("result", result_msg)

        # ── Write run log ──────────────────────────────────────────────────
        elapsed   = datetime.datetime.now() - run_start
        total_sec = int(elapsed.total_seconds())
        h, rem    = divmod(total_sec, 3600)
        m, s      = divmod(rem, 60)
        elapsed_str = f"{h}h {m:02d}m {s:02d}s" if h else f"{m:02d}m {s:02d}s"

        api_key = cfg.get("api_key", "")
        if len(api_key) > 8:
            api_masked = api_key[:4] + "*" * (len(api_key) - 8) + api_key[-4:]
        elif api_key:
            api_masked = "****"
        else:
            api_masked = "(not set)"

        status_str = "Stopped" if self._stop else "Completed"
        batches_ok = len(completed_batches)

        log_lines = [
            "BLAST Run Log",
            "=" * 60,
            f"Date/Time  : {run_start.strftime('%Y-%m-%d %H:%M:%S')}",
            f"Status     : {status_str}",
            f"Total time : {elapsed_str}",
            "",
            "Input files:",
        ]
        for f in self.files:
            log_lines.append(f"  {os.path.abspath(f)}")
        log_lines += [
            "",
            "Parameters:",
            f"  Database          : {cfg.get('database', '')}",
            f"  Program           : {cfg.get('program', '')}",
            f"  Hits per sequence : {nhits}",
            f"  Sequences / batch : {nseq}",
            f"  Fetch taxonomy    : {'Yes' if fetch_tax else 'No'}",
            f"  NCBI API key      : {api_masked}",
            "",
            "Results:",
            f"  Sequences found   : {seq_count}",
            f"  Batches           : {batches_ok}/{n_batches} completed",
            f"  Hits written      : {total_hits_done}",
            f"  Output folder     : {output_dir}",
            f"  TSV file          : {os.path.basename(tsv_path)}",
            f"  XLSX file         : {os.path.basename(xlsx_path) if xlsx_path else 'N/A'}",
        ]
        if miss_msg:
            log_lines.append(f"  Missing seqs      : {miss_msg}")
        log_lines += [
            "",
            "NOTE: Do not delete the .dbx cache files (accdb.dbx, taxadb.dbx).",
            "      They store organism and taxonomy lookups already performed and",
            "      will significantly speed up future BLAST runs on the same or",
            "      overlapping accession numbers.",
            "",
        ]

        os.makedirs(output_dir, exist_ok=True)
        log_path = os.path.join(output_dir, f"blast_run_log_{mydate}.txt")
        try:
            with open(log_path, "w", encoding="utf-8") as lf:
                lf.write("\n".join(log_lines))
        except Exception as exc:
            self.statusUpdated.emit("result", f"{result_msg}  │  Log error: {exc}")

        self.statusUpdated.emit(
            "info",
            f"Session stats  │  {self._req_count} requests · "
            f"{self._rate_limit_count} rate limit(s) (429) · "
            f"{self._server_err_count} server error(s) (5xx)"
        )
        self.taskFinished.emit(output_dir)
